#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Test amaçlı Excel dosyası oluşturucu
CATIA Automation Suite için örnek veri içerir
"""

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    print("⚠️ openpyxl bulunamadı. 'pip install openpyxl' ile yükleyin.")
    OPENPYXL_AVAILABLE = False
    exit(1)

def create_test_excel(output_path="test_data.xlsx"):
    """Test amaçlı Excel dosyası oluştur"""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl yüklü değil.")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Data"
    
    # Stil tanımlamaları
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="3B8ED0", end_color="3B8ED0", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    
    # Header satırı
    headers = ["ID", "Thickness", "Height", "Material", "X", "Y", "Z", "P1", "D1", "P2", "D2", "Angle", "Width", "Length"]
    ws.append(headers)
    
    # Header stilini uygula
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    
    # Test verileri (20 satır)
    test_data = [
        ["101", 5.0, 20.0, "AL", 10.5, 20.3, 30.1, 1.5, 0.5, 2.0, 0.8, 45.0, 100.0, 200.0],
        ["102", 5.2, 21.0, "AL", 12.0, 22.0, 32.0, 1.6, 0.6, 2.1, 0.9, 46.0, 101.0, 201.0],
        ["103", 5.4, 22.0, "ST", 14.0, 24.0, 34.0, 1.7, 0.7, 2.2, 1.0, 47.0, 102.0, 202.0],
        ["104", 5.6, 23.0, "ST", 16.0, 26.0, 36.0, 1.8, 0.8, 2.3, 1.1, 48.0, 103.0, 203.0],
        ["105", 5.8, 24.0, "AL", 18.0, 28.0, 38.0, 1.9, 0.9, 2.4, 1.2, 49.0, 104.0, 204.0],
        ["106", 6.0, 25.0, "AL", 20.0, 30.0, 40.0, 2.0, 1.0, 2.5, 1.3, 50.0, 105.0, 205.0],
        ["107", 6.2, 26.0, "ST", 22.0, 32.0, 42.0, 2.1, 1.1, 2.6, 1.4, 51.0, 106.0, 206.0],
        ["108", 6.4, 27.0, "ST", 24.0, 34.0, 44.0, 2.2, 1.2, 2.7, 1.5, 52.0, 107.0, 207.0],
        ["109", 6.6, 28.0, "AL", 26.0, 36.0, 46.0, 2.3, 1.3, 2.8, 1.6, 53.0, 108.0, 208.0],
        ["110", 6.8, 29.0, "AL", 28.0, 38.0, 48.0, 2.4, 1.4, 2.9, 1.7, 54.0, 109.0, 209.0],
        ["201", 7.0, 30.0, "ST", 30.0, 40.0, 50.0, 2.5, 1.5, 3.0, 1.8, 55.0, 110.0, 210.0],
        ["202", 7.2, 31.0, "ST", 32.0, 42.0, 52.0, 2.6, 1.6, 3.1, 1.9, 56.0, 111.0, 211.0],
        ["203", 7.4, 32.0, "AL", 34.0, 44.0, 54.0, 2.7, 1.7, 3.2, 2.0, 57.0, 112.0, 212.0],
        ["204", 7.6, 33.0, "AL", 36.0, 46.0, 56.0, 2.8, 1.8, 3.3, 2.1, 58.0, 113.0, 213.0],
        ["205", 7.8, 34.0, "ST", 38.0, 48.0, 58.0, 2.9, 1.9, 3.4, 2.2, 59.0, 114.0, 214.0],
        ["301", 8.0, 35.0, "ST", 40.0, 50.0, 60.0, 3.0, 2.0, 3.5, 2.3, 60.0, 115.0, 215.0],
        ["302", 8.2, 36.0, "AL", 42.0, 52.0, 62.0, 3.1, 2.1, 3.6, 2.4, 61.0, 116.0, 216.0],
        ["303", 8.4, 37.0, "AL", 44.0, 54.0, 64.0, 3.2, 2.2, 3.7, 2.5, 62.0, 117.0, 217.0],
        ["304", 8.6, 38.0, "ST", 46.0, 56.0, 66.0, 3.3, 2.3, 3.8, 2.6, 63.0, 118.0, 218.0],
        ["305", 8.8, 39.0, "ST", 48.0, 58.0, 68.0, 3.4, 2.4, 3.9, 2.7, 64.0, 119.0, 219.0],
    ]
    
    # Verileri ekle
    for row_data in test_data:
        ws.append(row_data)
    
    # Sütun genişliklerini ayarla
    column_widths = {
        'A': 12,  # ID
        'B': 10,  # Thickness
        'C': 10,  # Height
        'D': 12,  # Material
        'E': 8,   # X
        'F': 8,   # Y
        'G': 8,   # Z
        'H': 8,   # P1
        'I': 8,   # D1
        'J': 8,   # P2
        'K': 8,   # D2
        'L': 10,  # Angle
        'M': 10,  # Width
        'N': 10,  # Length
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Dosyayı kaydet
    wb.save(output_path)
    print(f"✅ Test Excel dosyası oluşturuldu: {output_path}")
    print(f"   - {len(test_data)} satır veri")
    print(f"   - {len(headers)} sütun")
    return output_path

if __name__ == "__main__":
    import sys
    output_file = sys.argv[1] if len(sys.argv) > 1 else "test_data.xlsx"
    create_test_excel(output_file)

