import customtkinter as ctk
import threading
import time
import datetime
import os
import sys
import string
import logging
import traceback
import re
from logging.handlers import RotatingFileHandler
from tkinter import filedialog, messagebox, Canvas
try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    TKDND_AVAILABLE = True
except ImportError:
    TKDND_AVAILABLE = False
    print("⚠️ tkinterdnd2 bulunamadı. Sürükle-bırak özelliği devre dışı. 'pip install tkinterdnd2' ile yükleyin.")
    # Dummy classes for compatibility
    class TkinterDnD:
        class DnDWrapper:
            pass
    DND_FILES = None

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️ openpyxl bulunamadı. 'pip install openpyxl' ile yükleyin (daha hızlı Excel okuma için)")
try:
    import win32com.client
    import pythoncom
    WIN32COM_AVAILABLE = True
except ImportError:
    WIN32COM_AVAILABLE = False
    print("⚠️ win32com bulunamadı. Test modu dışında çalışmayabilir.")

# ==========================================
# AYARLAR & YARDIMCILAR
# ==========================================
TEST_MODE = False  # True: Test Modu (Excel gerekmez) / False: Gerçek Mod

# Modern Renk Paleti (2025 Style)
THEME = {
    "bg_dark": "#0f172a",       # Çok koyu lacivert (Slate-900)
    "bg_card": "#1e293b",       # Koyu lacivert (Slate-800)
    "bg_card_hover": "#334155", # Slate-700
    "primary": "#3b82f6",       # Blue-500
    "primary_hover": "#2563eb", # Blue-600
    "secondary": "#8b5cf6",     # Violet-500
    "success": "#10b981",       # Emerald-500
    "danger": "#ef4444",        # Red-500
    "warning": "#f59e0b",       # Amber-500
    "text_main": "#f8fafc",     # Slate-50
    "text_muted": "#94a3b8",    # Slate-400
    "border": "#334155"         # Slate-700
}

# ==========================================
# MODERN WIDGET'LAR
# ==========================================
class ToastNotification(ctk.CTkToplevel):
    """Modern, kaybolan bildirim (Toast)"""
    def __init__(self, parent, message, icon="ℹ️", color=THEME["primary"], duration=3000):
        super().__init__(parent)
        self.overrideredirect(True)
        self.attributes("-topmost", True)
        # CTkToplevel için transparent fg_color desteklenmiyor, arka plan rengini kullan
        self.configure(fg_color=THEME["bg_dark"])
        
        # Ekran konumu (Sağ alt)
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        window_width = 320
        window_height = 70
        x = screen_width - window_width - 20
        y = screen_height - window_height - 60
        self.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        # Container (Rounded & Shadow effect simulation)
        self.frame = ctk.CTkFrame(self, fg_color=THEME["bg_card"], corner_radius=12, 
                                border_width=1, border_color=color)
        self.frame.pack(fill="both", expand=True)
        
        # İkon
        self.lbl_icon = ctk.CTkLabel(self.frame, text=icon, font=("Arial", 24))
        self.lbl_icon.pack(side="left", padx=(15, 10))
        
        # Mesaj
        self.lbl_msg = ctk.CTkLabel(self.frame, text=message, font=("Roboto", 13), 
                                   text_color=THEME["text_main"], wraplength=240, justify="left")
        self.lbl_msg.pack(side="left", padx=5, pady=10)
        
        # Kapatma butonu
        self.btn_close = ctk.CTkButton(self.frame, text="×", width=25, height=25,
                                      fg_color="transparent", hover_color=THEME["bg_card_hover"],
                                      text_color=THEME["text_muted"], font=("Arial", 16),
                                      command=self.destroy)
        self.btn_close.pack(side="right", padx=5, anchor="n", pady=5)
        
        # Animasyon ve otomatik kapanma
        self.alpha = 0.0
        self.attributes("-alpha", self.alpha)
        self.animate_in()
        self.after(duration, self.animate_out)
        
    def animate_in(self):
        if self.alpha < 1.0:
            self.alpha += 0.1
            self.attributes("-alpha", self.alpha)
            self.after(20, self.animate_in)
            
    def animate_out(self):
        if self.alpha > 0.0:
            self.alpha -= 0.1
            self.attributes("-alpha", self.alpha)
            self.after(20, self.animate_out)
        else:
            self.destroy()

class DropZone(ctk.CTkFrame):
    """Dosya Sürükle-Bırak Alanı"""
    def __init__(self, parent, command=None):
        super().__init__(parent, fg_color=THEME["bg_card"], corner_radius=15, 
                         border_width=2, border_color=THEME["border"])
        self.command = command
        self.bind("<Enter>", self.on_enter)
        self.bind("<Leave>", self.on_leave)
        self.bind("<Button-1>", self.on_click)
        
        # Not: Drop desteği ana pencereye eklendi, DropZone sadece görsel
        # Ana pencere zaten drop'u yakalayacak, burada sadece görsel feedback veriyoruz
        
        # İçerik
        self.inner_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.inner_frame.pack(expand=True, fill="both", padx=20, pady=20)
        
        self.icon_lbl = ctk.CTkLabel(self.inner_frame, text="📂", font=("Arial", 48))
        self.icon_lbl.pack(pady=(10, 5))
        
        self.text_lbl = ctk.CTkLabel(self.inner_frame, text="Dosyayı buraya sürükleyin\nveya seçmek için tıklayın",
                                    font=("Roboto", 14, "bold"), text_color=THEME["text_muted"])
        self.text_lbl.pack(pady=5)
        
        # Event binding (inner widgetlar için de)
        for widget in [self.inner_frame, self.icon_lbl, self.text_lbl]:
            widget.bind("<Enter>", self.on_enter)
            widget.bind("<Leave>", self.on_leave)
            widget.bind("<Button-1>", self.on_click)
    
    # DropZone'un kendi on_drop metodu yok - ana pencere zaten drop'u yakalıyor

    def on_enter(self, event):
        self.configure(border_color=THEME["primary"], fg_color=THEME["bg_card_hover"])
        # İkon rengini değiştir (scale desteklenmiyor)
        self.icon_lbl.configure(text_color=THEME["primary"])
        
    def on_leave(self, event):
        self.configure(border_color=THEME["border"], fg_color=THEME["bg_card"])
        self.icon_lbl.configure(text_color=THEME["text_muted"])
        
    def on_click(self, event):
        if self.command:
            self.command()


# ==========================================
# LOGGING SİSTEMİ
# ==========================================
def setup_logger():
    """Gelişmiş logging sistemi - Dosya ve konsol çıktısı"""
    if not os.path.exists("Logs"):
        os.makedirs("Logs")
    
    logger = logging.getLogger("CATIA_Automation")
    logger.setLevel(logging.DEBUG)
    
    # Eğer logger zaten handler'lara sahipse, tekrar ekleme
    if logger.handlers:
        return logger
    
    # Dosya handler (Rotating - Max 5MB, 5 yedek)
    log_file = os.path.join("Logs", "catia_automation.log")
    fh = RotatingFileHandler(log_file, maxBytes=5*1024*1024, backupCount=5, encoding='utf-8')
    fh.setLevel(logging.DEBUG)
    
    # Formatter
    formatter = logging.Formatter(
        '%(asctime)s - %(name)s - [%(levelname)s] - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    fh.setFormatter(formatter)
    
    logger.addHandler(fh)
    return logger

# Global logger
APP_LOGGER = setup_logger()

# ==========================================
# VERİ DOĞRULAMA
# ==========================================
def validate_parameter_value(value, param_name):
    """Parametre değerini CATIA'ya göndermeden önce doğrula"""
    try:
        float_val = float(value)
        
        # Özel doğrulama kuralları
        if "Thickness" in param_name or param_name.startswith("T"):
            if float_val <= 0:
                raise ValueError(f"{param_name} pozitif olmalı")
        
        if "Angle" in param_name:
            if not -360 <= float_val <= 360:
                raise ValueError(f"{param_name} -360 ile 360 arasında olmalı")
        
        return float_val
    except ValueError as e:
        raise ValueError(f"{param_name} için geçersiz değer '{value}': {str(e)}")

# ==========================================
# EXCEL İŞLEMLERİ
# ==========================================
def read_excel_openpyxl(file_path, sheet_name=None):
    """openpyxl ile hızlı Excel okuma"""
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        # Tüm satırları oku (ilk satır header)
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:  # İlk sütun (ID) boş değilse
                data.append(row)
        
        # Sayfa isimlerini al
        sheets = wb.sheetnames
        wb.close()
        
        return data, sheets
    except Exception as e:
        APP_LOGGER.error(f"openpyxl okuma hatası: {e}")
        raise

def read_excel_preview_openpyxl(file_path, max_rows=10):
    """openpyxl ile önizleme okuma (ilk N satır, tüm sütunlar)"""
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        
        # En son dolu sütunu bul - önce tüm satırları okuyup en uzun satırı bul
        max_col = 0
        preview_rows = min(max_rows, ws.max_row) if ws.max_row else max_rows
        
        # İlk önce tüm satırları oku ve en uzun satırı bul
        temp_data = []
        for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=preview_rows, values_only=True)):
            if idx >= max_rows:
                break
            row_list = list(row) if row else []
            # None değerlerini temizle ve gerçek uzunluğu bul
            # Sondaki None'ları temizle
            while row_list and row_list[-1] is None:
                row_list.pop()
            if len(row_list) > max_col:
                max_col = len(row_list)
            temp_data.append(row_list)
        
        # Eğer hiç veri yoksa, en azından A sütununu oku
        if max_col == 0:
            max_col = 1
        
        # Şimdi tüm satırları aynı uzunluğa tamamla
        data = []
        for row_list in temp_data:
            # Satırı maksimum sütun sayısına tamamla
            while len(row_list) < max_col:
                row_list.append(None)
            data.append(row_list)
        
        sheets = wb.sheetnames
        wb.close()
        
        return data, sheets
    except Exception as e:
        APP_LOGGER.error(f"openpyxl önizleme hatası: {e}")
        raise

def col2num(col_str):
    """Harfi sayıya çevirir (A->1, Z->26, AA->27)"""
    num = 0
    col_str = col_str.strip().upper()
    for c in col_str:
        if c in string.ascii_letters:
            num = num * 26 + (ord(c) - ord('A')) + 1
    return num

def num2col(n):
    """Sayıyı harfe çevirir (1->A, 27->AA)"""
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string

# ==========================================
# ARKA PLAN İŞÇİSİ (DATA OKUMA & YAZMA)
# ==========================================
class WorkerThread(threading.Thread):
    def __init__(self, app, excel_path, config, dynamic_params):
        super().__init__()
        self.app = app
        self.excel_path = excel_path
        self.config = config
        self.dynamic_params = dynamic_params # List of tuples: (Suffix, ColChar)
        self.running = True
        self.daemon = True

    def run(self):
        if TEST_MODE:
            self.run_simulation()
        else:
            self.run_real_process()

    def run_simulation(self):
        sheet_name = self.config.get("sheet_name", "Sheet1")
        self.app.log(f"Başladı: {sheet_name}", "info")
        
        # Dinamik parametreleri logla
        msg = "Parametreler: "
        for suffix, col in self.dynamic_params:
            msg += f"[{suffix}->{col}] "
        self.app.log(msg, "info")
        
        time.sleep(1)
        total = 50
        self.app.after(0, self.app.update_max_progress, total)
        
        updates = 0
        errors = 0
        
        for i in range(1, total + 1):
            if not self.running: break
            time.sleep(0.05)
            # Progress güncelle
            if hasattr(self.app, 'progress_bar'):
                self.app.after(0, lambda x=i, t=total: self.app.progress_bar.set(x / t))
            if hasattr(self.app, 'lbl_progress'):
                self.app.after(0, lambda x=i, t=total: self.app.lbl_progress.configure(text=f"{x} / {t}"))
            # İstatistikleri güncelle
            if hasattr(self.app, 'card_success') and hasattr(self.app, 'card_error'):
                self.app.after(0, lambda u=updates, e=errors: (
                    self.app.card_success.configure(text=str(u)),
                    self.app.card_error.configure(text=str(e))
                ))
            
            if i % 5 == 0:
                updates += 1
                # Örnek: İlk dinamik parametreyi update etmiş gibi yapalım
                p_name = f"Rib_{i}_{self.dynamic_params[0][0]}" if self.dynamic_params else "Param"
                self.app.after(0, self.app.log, f"{p_name} güncellendi.", "update")

        self.app.after(0, self.app.finish_process)

    def run_real_process(self):
        """Gerçek işlem - Excel okuma ve CATIA yazma"""
        excel = None
        catia = None
        
        try:
            APP_LOGGER.info("İşlem başlatıldı")
            
            # openpyxl ile okuma (varsa, daha hızlı)
            if OPENPYXL_AVAILABLE:
                APP_LOGGER.info("openpyxl ile Excel okunuyor...")
                self.run_with_openpyxl()
                return
            
            # Fallback: win32com ile okuma
            if not WIN32COM_AVAILABLE:
                raise ImportError("win32com bulunamadı. 'pip install pywin32' ile yükleyin.")
            
            APP_LOGGER.info("win32com ile Excel okunuyor...")
            import win32com.client
            import pythoncom
            pythoncom.CoInitialize()

            # Excel'i görünmez modda aç ve performans ayarları
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False  # Görünmez mod - daha hızlı
            excel.ScreenUpdating = False  # Ekran güncellemelerini kapat
            excel.DisplayAlerts = False  # Uyarıları kapat
            excel.EnableEvents = False  # Event'leri kapat
            excel.Calculation = -4135  # xlCalculationManual (otomatik hesaplamayı kapat)
            
            wb = excel.Workbooks.Open(self.excel_path, ReadOnly=True)  # Read-only aç - daha hızlı
            target_sheet = self.config.get("sheet_name", "")
            
            # Sayfa bul
            try: valid_sheet = wb.Sheets(target_sheet)
            except: valid_sheet = wb.ActiveSheet
                
            last_row = valid_sheet.Cells(valid_sheet.Rows.Count, 1).End(-4162).Row
            
            # Sadece gerekli sütunları oku (optimizasyon)
            # En sağdaki sütunu bul
            max_col_idx = 0
            for suffix, col_letter in self.dynamic_params:
                if col_letter:
                    col_idx = col2num(col_letter)
                    if col_idx > max_col_idx:
                        max_col_idx = col_idx
            
            # A sütunu (ID) + en sağdaki sütun + biraz buffer
            end_col = num2col(max(max_col_idx, 26))  # En az Z, gerekirse daha fazla
            
            # Excel verisini komple belleğe al (Hız için)
            raw_data = valid_sheet.Range(f"A2:{end_col}{last_row}").Value 
            # raw_data tuple of tuples döner. raw_data[satir_idx][sutun_idx]
            
            total_rows = len(raw_data) if raw_data else 0
            self.app.after(0, self.app.update_max_progress, total_rows)
            
            updates = 0
            errors = 0
            
            # --- PARAMETRE MAP --- (önceden hesapla)
            # suffix boş olsa bile col_letter varsa ekle (Thickness gibi durumlar için)
            param_map = []
            for suffix, col_letter in self.dynamic_params:
                if not col_letter: continue
                col_idx = col2num(col_letter) - 1  # Python 0-based index
                # suffix boş string olsa bile ekle (sadece ID kullanılacak)
                param_map.append((suffix if suffix else "", col_idx))
            
            # CATIA bağlantısı (eğer gerekiyorsa)
            catia = None
            try:
                catia = win32com.client.GetActiveObject("CATIA.Application")
                # CATIA optimizasyonları
                if catia:
                    # Screen refresh'i kapat (batch update için)
                    # catia.DisplayFileAlerts = False
                    pass
            except:
                pass  # CATIA yoksa devam et
            
            # Batch update için değişiklikleri topla
            batch_size = 50  # Her 50 satırda bir UI güncelle
            last_update = 0
            
            # DÖNGÜ (optimize edilmiş)
            if raw_data:
                # Tuple yerine list'e çevir (daha hızlı erişim)
                data_list = list(raw_data) if isinstance(raw_data, tuple) else raw_data
                
                for i, row in enumerate(data_list):
                    if not self.running: break
                    
                    # ID Okuma (A Sütunu -> index 0)
                    try: 
                        id_val = row[0] if len(row) > 0 else None
                    except: 
                        continue
                    
                    if not id_val: continue
                    
                    # ID Formatlama (optimize edilmiş)
                    if isinstance(id_val, float) and id_val.is_integer():
                        id_str = str(int(id_val))
                    else:
                        id_str = str(id_val).strip()
                    
                    if not id_str: continue
                    
                    # Dinamik Parametreleri Güncelle
                    row_updates = 0
                    for suffix, col_idx in param_map:
                        if col_idx < len(row):
                            val = row[col_idx]
                            if val is not None and val != "":
                                # CATIA Parametre Adı: ID + Suffix (suffix boş olsa bile ID kullan)
                                # Eğer suffix boşsa, sadece ID kullanılır (örn: "10" yerine "10Thickness" değil "10")
                                full_name = id_str + (suffix if suffix else "")
                                
                                # --- CATIA YAZMA KODU ---
                                try:
                                    # Değeri doğrula
                                    validated_value = validate_parameter_value(val, full_name)
                                    
                                    if catia:
                                        # CATIA'ya parametre yaz
                                        try:
                                            doc = catia.ActiveDocument
                                            part = doc.Part
                                            param = part.Parameters.Item(full_name)
                                            param.Value = validated_value
                                            # Parametreyi güncelle
                                            part.Update()
                                            APP_LOGGER.debug(f"{full_name} = {validated_value}")
                                        except Exception as catia_err:
                                            APP_LOGGER.error(f"CATIA parametresi yazılamadı ({full_name}): {catia_err}")
                                            raise
                                    
                                    updates += 1
                                    row_updates += 1
                                except ValueError as ve:
                                    errors += 1
                                    APP_LOGGER.warning(f"Doğrulama hatası - Satır {i+2}: {ve}")
                                    if errors <= 10:
                                        self.app.after(0, self.app.log, f"Satır {i+2}: {str(ve)}", "error")
                                except Exception as e:
                                    errors += 1
                                    APP_LOGGER.error(f"Beklenmeyen hata - Satır {i+2}: {e}\n{traceback.format_exc()}")
                                    if errors <= 10:
                                        self.app.after(0, self.app.log, f"Satır {i+2}: {full_name} = {str(e)}", "error")
                    
                    # Batch UI güncelleme (her N satırda bir veya son satır)
                    if (i + 1 - last_update >= batch_size) or (i + 1 == total_rows):
                        self.app.after(0, self.app.update_stats, i+1, updates, errors)
                        last_update = i + 1

            # Excel'i kapat
            wb.Close(False)
            excel.Quit()
            
            # Son güncelleme
            self.app.after(0, self.app.update_stats, total_rows, updates, errors)
            self.app.after(0, self.app.finish_process)

        except Exception as e:
            APP_LOGGER.critical(f"Kritik hata (real_process): {e}\n{traceback.format_exc()}")
            self.app.after(0, self.app.log, f"KRİTİK HATA: {e}", "error")
            
            # Cleanup - Excel'i güvenli şekilde kapat
            try:
                if 'excel' in locals() and excel:
                    APP_LOGGER.info("Excel kapatılıyor...")
                    if 'wb' in locals():
                        wb.Close(False)
                    excel.Quit()
                    APP_LOGGER.info("Excel kapatıldı")
            except Exception as cleanup_err:
                APP_LOGGER.error(f"Excel cleanup hatası: {cleanup_err}")
            
            # CATIA'yı serbest bırak
            try:
                if 'catia' in locals() and catia:
                    del catia
            except:
                pass
            
            # İşlemi bitir
            self.app.after(0, self.app.finish_process)

    def run_with_openpyxl(self):
        """openpyxl ile hızlı Excel okuma ve işleme"""
        try:
            APP_LOGGER.info(f"openpyxl ile dosya açılıyor: {self.excel_path}")
            sheet_name = self.config.get("sheet_name", None)
            
            # Excel'i oku
            data, sheets = read_excel_openpyxl(self.excel_path, sheet_name)
            
            total_rows = len(data)
            self.app.after(0, self.app.update_max_progress, total_rows)
            APP_LOGGER.info(f"Toplam {total_rows} satır okundu")
            
            updates = 0
            errors = 0
            
            # Parametre mapping'i hazırla
            param_map = []
            for suffix, col_letter in self.dynamic_params:
                if not col_letter:
                    continue
                col_idx = col2num(col_letter) - 1  # Python 0-based index
                param_map.append((suffix, col_idx))
            
            APP_LOGGER.info(f"Parametre mapping: {param_map}")
            
            # CATIA bağlantısı
            catia = None
            try:
                if WIN32COM_AVAILABLE:
                    import win32com.client
                    catia = win32com.client.GetActiveObject("CATIA.Application")
                    APP_LOGGER.info("CATIA bağlantısı başarılı")
                else:
                    APP_LOGGER.warning("win32com yok, CATIA'ya yazılamayacak (sadece simülasyon)")
            except Exception as catia_err:
                APP_LOGGER.warning(f"CATIA bağlanamadı: {catia_err}")
                self.app.after(0, self.app.log, "CATIA bağlanamadı - sadece simülasyon modu", "error")
            
            # Batch güncelleme
            batch_size = 50
            last_update = 0
            
            # Satırları işle
            for i, row in enumerate(data):
                if not self.running:
                    APP_LOGGER.info("İşlem kullanıcı tarafından durduruldu")
                    break
                
                # ID oku (ilk sütun)
                id_val = row[0] if len(row) > 0 else None
                if not id_val:
                    continue
                
                # ID formatla
                if isinstance(id_val, (int, float)):
                    id_str = str(int(id_val)) if float(id_val).is_integer() else str(id_val)
                else:
                    id_str = str(id_val).strip()
                
                if not id_str:
                    continue
                
                # Parametreleri güncelle
                row_updates = 0
                for suffix, col_idx in param_map:
                    if col_idx < len(row):
                        val = row[col_idx]
                        if val is not None and val != "":
                            full_name = id_str + suffix
                            
                            try:
                                # Değeri doğrula
                                validated_value = validate_parameter_value(val, full_name)
                                
                                if catia:
                                    # CATIA'ya yaz
                                    try:
                                        doc = catia.ActiveDocument
                                        part = doc.Part
                                        param = part.Parameters.Item(full_name)
                                        param.Value = validated_value
                                        part.Update()
                                        APP_LOGGER.debug(f"{full_name} = {validated_value}")
                                    except Exception as catia_err:
                                        APP_LOGGER.error(f"CATIA yazma hatası ({full_name}): {catia_err}")
                                        raise
                                
                                updates += 1
                                row_updates += 1
                                
                            except ValueError as ve:
                                errors += 1
                                APP_LOGGER.warning(f"Doğrulama hatası - Satır {i+2}: {ve}")
                                if errors <= 10:
                                    self.app.after(0, self.app.log, f"Satır {i+2}: {str(ve)}", "error")
                            except Exception as e:
                                errors += 1
                                APP_LOGGER.error(f"Hata - Satır {i+2}: {e}\n{traceback.format_exc()}")
                                if errors <= 10:
                                    self.app.after(0, self.app.log, f"Satır {i+2}: {full_name} = {str(e)}", "error")
                
                # Batch UI güncelleme
                if (i + 1 - last_update >= batch_size) or (i + 1 == total_rows):
                    self.app.after(0, self.app.update_stats, i+1, updates, errors)
                    last_update = i + 1
            
            # Son güncelleme
            self.app.after(0, self.app.update_stats, total_rows, updates, errors)
            self.app.after(0, self.app.finish_process)
            APP_LOGGER.info(f"İşlem tamamlandı - Başarılı: {updates}, Hata: {errors}")
            
        except Exception as e:
            APP_LOGGER.error(f"Kritik hata (openpyxl): {e}\n{traceback.format_exc()}")
            self.app.after(0, self.app.log, f"KRİTİK HATA: {e}", "error")
            self.app.after(0, self.app.finish_process)
    
    def stop(self):
        self.running = False
        APP_LOGGER.info("Durdurma talebi alındı")

# ==========================================
# EXCEL PREVIEW & ANALİZ
# ==========================================
class ExcelPreviewLoader(threading.Thread):
    def __init__(self, app, path):
        super().__init__()
        self.app = app
        self.path = path
        self.daemon = True

    def run(self):
        try:
            data_preview = []
            sheets = []
            
            if TEST_MODE:
                time.sleep(0.5)
                sheets = ["Visualization Data", "Sheet2"]
                # Fake Data: Header + 5 Rows
                data_preview = [
                    ["ID", "Th", "Height", "Mat", "X", "Y", "Z", "Q", "W", "E", "P1", "D1", "P2", "D2"],
                    ["101", "5.0", "20.0", "AL", "10", "20", "30", "-", "-", "-", "1.5", "0.5", "2.0", "0.8"],
                    ["102", "5.2", "21.0", "AL", "12", "22", "32", "-", "-", "-", "1.6", "0.6", "2.1", "0.9"],
                    ["103", "5.4", "22.0", "ST", "14", "24", "34", "-", "-", "-", "1.7", "0.7", "2.2", "1.0"],
                    ["104", "5.6", "23.0", "ST", "16", "26", "36", "-", "-", "-", "1.8", "0.8", "2.3", "1.1"],
                ]
            elif OPENPYXL_AVAILABLE:
                # openpyxl ile önizleme (daha hızlı)
                APP_LOGGER.info("openpyxl ile önizleme yükleniyor...")
                data_preview, sheets = read_excel_preview_openpyxl(self.path, max_rows=10)
            elif WIN32COM_AVAILABLE:
                # Fallback: win32com ile önizleme
                APP_LOGGER.info("win32com ile önizleme yükleniyor...")
                import win32com.client
                import pythoncom
                pythoncom.CoInitialize()
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                excel.ScreenUpdating = False
                excel.DisplayAlerts = False
                excel.EnableEvents = False
                excel.Calculation = -4135  # xlCalculationManual
                
                wb = excel.Workbooks.Open(self.path, ReadOnly=True)
                
                # Sayfaları al
                sheets = [s.Name for s in wb.Sheets]
                
                # İlk sayfadan önizleme al
                ws = wb.Sheets(1)
                
                # En son dolu satır ve sütunu bul
                last_row = ws.Cells(ws.Rows.Count, 1).End(-4162).Row  # xlUp = -4162
                last_col = ws.Cells(1, ws.Columns.Count).End(-4159).Column  # xlToLeft = -4159
                
                # Maksimum 10 satır ve tüm sütunları oku
                max_preview_rows = min(10, last_row)
                end_col_letter = num2col(last_col)
                
                # Range oluştur (A1:SonSütun10)
                range_str = f"A1:{end_col_letter}{max_preview_rows}"
                vals = ws.Range(range_str).Value
                
                # Tuple to List (batch conversion)
                if vals:
                    if isinstance(vals, tuple):
                        data_preview = [list(row) if row else [] for row in vals]
                    else:
                        data_preview = [list(row) if row else [] for row in [vals]]
                
                wb.Close(False)
                excel.Quit()
            else:
                raise ImportError("Excel okuma için openpyxl veya pywin32 gerekli")

            self.app.after(0, self.app.update_ui_with_excel_data, sheets, data_preview)

        except Exception as e:
            APP_LOGGER.error(f"Önizleme hatası: {e}\n{traceback.format_exc()}")
            self.app.after(0, self.app.log, f"Önizleme hatası: {e}", "error")
            
            # Hata detayını göster
            error_msg = f"Excel önizlemesi yüklenemedi:\n\n{str(e)}\n\n"
            if not OPENPYXL_AVAILABLE and not WIN32COM_AVAILABLE:
                error_msg += "Çözüm: 'pip install openpyxl' veya 'pip install pywin32' ile gerekli kütüphaneleri yükleyin."
            
            self.app.after(0, messagebox.showerror, "Önizleme Hatası", error_msg)

# ==========================================
# ÖZEL BUTON WIDGET'LARI
# ==========================================
# ==========================================
# ARAYÜZ (GUI)
# ==========================================
class AutomationSuite(ctk.CTk):
    def __init__(self):
        super().__init__()
        
        # Drag & Drop desteği (opsiyonel)
        if TKDND_AVAILABLE:
            try:
                self.TkdndVersion = TkinterDnD._require(self)
                self.drop_target_register(DND_FILES)
                self.dnd_bind('<<Drop>>', self.on_drop)
                APP_LOGGER.info("TkinterDnD başarıyla yüklendi")
            except Exception as e:
                APP_LOGGER.error(f"TkinterDnD yükleme hatası: {e}\n{traceback.format_exc()}")
        
        APP_LOGGER.info("Uygulama başlatılıyor...")
        
        # Config
        self.config = {"sheet_name": "", "always_on_top": False}
        self.param_rows = [] 

        # Tema ve Pencere
        ctk.set_appearance_mode("Dark")
        ctk.set_default_color_theme("dark-blue")
        
        self.title("AFT Sizing Automation")
        self.geometry("1200x800")
        self.configure(fg_color=THEME["bg_dark"])
        
        # Uygulama adı ve icon (tüm platformlar için)
        # macOS için özel ayarlar
        if sys.platform == "darwin":
            try:
                from AppKit import NSApplication
                app = NSApplication.sharedApplication()
                app.setApplicationName_("AFT Sizing Automation")
            except ImportError:
                try:
                    self.tk.call('wm', 'iconname', self._w, 'AFT Sizing Automation')
                except:
                    pass
            except:
                pass
        
        # Windows için uygulama adı
        if sys.platform == "win32":
            try:
                import ctypes
                ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("AFT.Sizing.Automation")
            except:
                pass
        
        # Icon oluştur (tüm platformlar için)
        try:
            from PIL import Image, ImageDraw, ImageFont, ImageTk
            # 64x64 icon oluştur (daha yüksek kalite)
            icon = Image.new('RGBA', (64, 64), (0, 0, 0, 0))
            draw = ImageDraw.Draw(icon)
            
            # Modern gradient arka plan (mavi tonları)
            # Üst kısım - açık mavi
            draw.rectangle([0, 0, 64, 32], fill=(59, 130, 246, 255))  # Blue-500
            # Alt kısım - koyu mavi
            draw.rectangle([0, 32, 64, 64], fill=(37, 99, 235, 255))  # Blue-600
            
            # "A" harfi için modern tasarım
            # Üçgen şeklinde "A" harfi
            # Sol üst çizgi
            draw.line([20, 50, 32, 20], fill=(255, 255, 255, 255), width=4)
            # Sağ üst çizgi
            draw.line([44, 50, 32, 20], fill=(255, 255, 255, 255), width=4)
            # Yatay çizgi (A'nın ortası)
            draw.line([26, 38, 38, 38], fill=(255, 255, 255, 255), width=3)
            
            # Font ile "A" harfi ekle (daha net görünüm için)
            font = None
            if sys.platform == "darwin":
                try:
                    font = ImageFont.truetype("/System/Library/Fonts/Helvetica.ttc", 42)
                except:
                    try:
                        font = ImageFont.truetype("/System/Library/Fonts/HelveticaNeue.ttc", 42)
                    except:
                        pass
            elif sys.platform == "win32":
                try:
                    font = ImageFont.truetype("C:/Windows/Fonts/arial.ttf", 42)
                except:
                    pass
            
            if font:
                # "A" harfini kalın ve beyaz renkte çiz
                draw.text((18, 10), "A", fill=(255, 255, 255, 255), font=font)
            else:
                # Font yoksa daha kalın çizgilerle "A" çiz
                draw.line([22, 48, 32, 18], fill=(255, 255, 255, 255), width=5)
                draw.line([42, 48, 32, 18], fill=(255, 255, 255, 255), width=5)
                draw.line([26, 36, 38, 36], fill=(255, 255, 255, 255), width=4)
            
            # Icon'u kullan
            photo = ImageTk.PhotoImage(icon)
            self.iconphoto(False, photo)
            # Referansı sakla (garbage collection'ı önlemek için)
            self._icon_photo = photo
        except ImportError:
            APP_LOGGER.debug("PIL yüklü değil, icon atlanıyor")
        except Exception as e:
            APP_LOGGER.debug(f"Icon oluşturulamadı: {e}")
        
        # Değişkenler
        self.log_entries = []
        self.show_errors_only = False
        self.preview_table = None
        self.tab_titles = {"monitor": "  🚀 Monitör  ", "settings": "  ⚙️ Ayarlar  "}
        self.current_run_has_error = False
        self.selected_file = None
        self.worker = None
        self.total_work = 1
        self.start_time = 0
        
        # --- ANA LAYOUT ---
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)
        
        # TabView (Modern Style)
        self.tab_view = ctk.CTkTabview(self, corner_radius=20, fg_color=THEME["bg_dark"],
                                      segmented_button_fg_color=THEME["bg_card"],
                                      segmented_button_selected_color=THEME["primary"],
                                      segmented_button_selected_hover_color=THEME["primary_hover"],
                                      segmented_button_unselected_color=THEME["bg_card"],
                                      segmented_button_unselected_hover_color=THEME["bg_card_hover"],
                                      text_color=THEME["text_main"])
        self.tab_view.grid(row=0, column=0, padx=20, pady=20, sticky="nsew")
        
        self.tab_monitor = self.tab_view.add("  🚀 Monitör  ")
        self.tab_settings = self.tab_view.add("  ⚙️ Ayarlar  ")
        
        self.setup_monitor()
        self.setup_settings()
        
        # Klavye kısayolları
        self.bind("<Control-r>", lambda e: self.start_process() if self.btn_run.cget("state") == "normal" else None)
        self.bind("<Control-w>", lambda e: self.stop_process() if self.btn_stop.cget("state") == "normal" else None)
        self.bind("<Control-f>", lambda e: self.select_file())
        
        self.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        # Pencereyi ekranın önüne getir
        self.lift()
        self.attributes("-topmost", True)
        self.after(100, lambda: self.attributes("-topmost", False))  # 100ms sonra topmost'u kapat
        self.focus_force()

    def on_drop(self, event):
        """Sürükle-bırak dosya yükleme (opsiyonel)"""
        if not TKDND_AVAILABLE:
            self.show_toast("Uyarı", "Sürükle-bırak desteği yüklü değil.", type="warning")
            return
        
        try:
            APP_LOGGER.info(f"Drop event alındı: {type(event)}")
            
            # Event.data'yı al
            if not hasattr(event, 'data'):
                APP_LOGGER.error("Drop event'inde 'data' attribute'u yok")
                self.show_toast("Hata", "Drop event'i geçersiz.", type="error")
                return
            
            # Dosya listesini al
            data_str = str(event.data)
            APP_LOGGER.debug(f"Raw drop data: {data_str[:200]}...")  # İlk 200 karakter
            
            # macOS için özel işleme
            if sys.platform == "darwin":
                # macOS'ta format: {file:///path/to/file} veya file:///path/to/file
                # Önce { } karakterlerini temizle
                data_str = data_str.strip('{}')
                # file:// prefix'ini temizle
                if data_str.startswith('file://'):
                    data_str = data_str[7:]
                # URL decode gerekebilir (%20 -> space)
                try:
                    from urllib.parse import unquote
                    data_str = unquote(data_str)
                except:
                    pass
                files = [data_str]
            else:
                # Windows/Linux için normal splitlist
                try:
                    files = self.tk.splitlist(event.data)
                except:
                    # Fallback: string olarak al
                    files = [str(event.data)]
            
            APP_LOGGER.debug(f"Parsed files count: {len(files)}")
            
            if files:
                path = files[0]
                
                # Path temizleme
                path = path.strip()
                # file:// prefix'ini temizle (eğer hala varsa)
                if path.startswith('file://'):
                    path = path[7:]
                # { } karakterlerini temizle
                path = path.strip('{}')
                # URL decode tekrar (eğer gerekirse)
                try:
                    from urllib.parse import unquote
                    path = unquote(path)
                except:
                    pass
                
                # Absolute path kontrolü
                if not os.path.isabs(path):
                    path = os.path.abspath(path)
                
                APP_LOGGER.info(f"Drop ile dosya seçildi: {path}")
                
                if self.validate_excel_file(path):
                    self.load_excel_file(path)
                    self.show_toast("Başarılı", f"Dosya yüklendi:\n{os.path.basename(path)}", type="success")
                else:
                    self.show_toast("Geçersiz Dosya", "Lütfen bir Excel dosyası sürükleyin.", type="error")
            else:
                APP_LOGGER.warning("Drop event'inde dosya bulunamadı")
                self.show_toast("Hata", "Dosya algılanamadı.", type="error")
        except Exception as e:
            APP_LOGGER.error(f"Drop hatası: {e}\n{traceback.format_exc()}")
            self.show_toast("Hata", f"Dosya yüklenemedi:\n{str(e)}", type="error")

    def show_toast(self, title, message, type="info"):
        color = THEME["primary"]
        icon = "ℹ️"
        if type == "success": 
            color = THEME["success"]
            icon = "✅"
        elif type == "error": 
            color = THEME["danger"]
            icon = "❌"
        elif type == "warning":
            color = THEME["warning"]
            icon = "⚠️"
            
        ToastNotification(self, f"{title}\n{message}", icon, color)

    # ------------------------------
    # MONİTÖR TABI (YENİ KART TASARIMI)
    # ------------------------------
    def setup_monitor(self):
        self.tab_monitor.grid_columnconfigure(0, weight=1)
        self.tab_monitor.grid_columnconfigure(1, weight=2) # Sağ taraf daha geniş
        self.tab_monitor.grid_rowconfigure(0, weight=1)
        
        # --- SOL KOLON (Kontrol & Dosya) ---
        left_col = ctk.CTkFrame(self.tab_monitor, fg_color="transparent")
        left_col.grid(row=0, column=0, padx=(0, 10), pady=10, sticky="nsew")
        
        # 1. Dosya Kartı
        file_card = ctk.CTkFrame(left_col, fg_color=THEME["bg_card"], corner_radius=15)
        file_card.pack(fill="x", pady=(0, 15))
        
        ctk.CTkLabel(file_card, text="Dosya Seçimi", font=("Roboto", 16, "bold"), text_color=THEME["text_main"]).pack(anchor="w", padx=20, pady=(15, 10))
        
        self.drop_zone = DropZone(file_card, command=self.select_file)
        self.drop_zone.pack(fill="x", padx=20, pady=(0, 20))
        
        # DropZone ve file_card'a da drop desteği ekle (macOS için)
        if TKDND_AVAILABLE:
            try:
                # file_card'a drop desteği ekle
                file_card.drop_target_register(DND_FILES)
                file_card.dnd_bind('<<Drop>>', self.on_drop)
                # DropZone'un alt widget'larına da ekle
                if hasattr(self.drop_zone, '_canvas'):
                    self.drop_zone._canvas.drop_target_register(DND_FILES)
                    self.drop_zone._canvas.dnd_bind('<<Drop>>', self.on_drop)
            except Exception as e:
                APP_LOGGER.debug(f"file_card drop desteği eklenemedi: {e}")
        
        # Dosya bilgisi ve silme butonu için frame
        file_info_frame = ctk.CTkFrame(file_card, fg_color="transparent")
        file_info_frame.pack(fill="x", padx=20, pady=(0, 15))
        
        self.lbl_file_info = ctk.CTkLabel(file_info_frame, text="Dosya seçilmedi", 
                                         font=("Roboto", 12), text_color=THEME["text_muted"])
        self.lbl_file_info.pack(side="left", fill="x", expand=True)
        
        self.btn_remove_file = ctk.CTkButton(file_info_frame, text="🗑️", 
                                            command=self.remove_file,
                                            width=35, height=35,
                                            fg_color="transparent",
                                            hover_color=THEME["danger"],
                                            text_color=THEME["text_muted"],
                                            font=("Arial", 14),
                                            state="disabled")
        self.btn_remove_file.pack(side="right", padx=(10, 0))

        # 2. İşlem Kartı
        action_card = ctk.CTkFrame(left_col, fg_color=THEME["bg_card"], corner_radius=15)
        action_card.pack(fill="x", pady=(0, 15))
        
        ctk.CTkLabel(action_card, text="İşlem Kontrolü", font=("Roboto", 16, "bold"), text_color=THEME["text_main"]).pack(anchor="w", padx=20, pady=(15, 10))
        
        self.btn_run = ctk.CTkButton(action_card, text="🚀 BAŞLAT", command=self.start_process,
                                    fg_color=THEME["success"], hover_color="#059669",
                                    height=50, font=("Roboto", 16, "bold"), corner_radius=10)
        self.btn_run.pack(fill="x", padx=20, pady=(0, 10))
        
        self.btn_stop = ctk.CTkButton(action_card, text="🛑 DURDUR", command=self.stop_process,
                                     fg_color=THEME["danger"], hover_color="#dc2626",
                                     height=40, font=("Roboto", 14, "bold"), corner_radius=10, state="disabled")
        self.btn_stop.pack(fill="x", padx=20, pady=(0, 20))

        # --- SAĞ KOLON (İstatistik & Log) ---
        right_col = ctk.CTkFrame(self.tab_monitor, fg_color="transparent")
        right_col.grid(row=0, column=1, sticky="nsew")
        right_col.rowconfigure(2, weight=1) # Log alanı esnek
        
        # 1. İstatistik Kartları (Grid)
        stats_grid = ctk.CTkFrame(right_col, fg_color="transparent")
        stats_grid.pack(fill="x", pady=(0, 15))
        stats_grid.columnconfigure((0,1,2), weight=1)
        
        self.card_success = self._stat_card(stats_grid, "Başarılı", "0", THEME["success"], 0)
        self.card_error = self._stat_card(stats_grid, "Hatalar", "0", THEME["danger"], 1)
        self.card_time = self._stat_card(stats_grid, "Geçen Süre", "00:00", THEME["primary"], 2)
        
        # 2. Progress Bar Kartı (İşlem Hızı yerine)
        progress_card = ctk.CTkFrame(right_col, fg_color=THEME["bg_card"], corner_radius=15)
        progress_card.pack(fill="x", pady=(0, 15))
        
        header_frame = ctk.CTkFrame(progress_card, fg_color="transparent")
        header_frame.pack(fill="x", padx=20, pady=(15, 10))
        ctk.CTkLabel(header_frame, text="📊 İlerleme", font=("Roboto", 14, "bold")).pack(side="left")
        self.lbl_progress = ctk.CTkLabel(header_frame, text="0 / 0", text_color=THEME["text_muted"], font=("Roboto", 13))
        self.lbl_progress.pack(side="right")
        
        self.progress_bar = ctk.CTkProgressBar(progress_card, height=25, corner_radius=12,
                                              progress_color=THEME["primary"], 
                                              fg_color=THEME["bg_dark"],
                                              border_width=0)
        self.progress_bar.pack(fill="x", padx=20, pady=(0, 20))
        self.progress_bar.set(0)
        
        # 3. Terminal/Log Kartı
        log_card = ctk.CTkFrame(right_col, fg_color=THEME["bg_card"], corner_radius=15)
        log_card.pack(fill="both", expand=True)
        
        log_header = ctk.CTkFrame(log_card, fg_color="transparent")
        log_header.pack(fill="x", padx=20, pady=10)
        ctk.CTkLabel(log_header, text="Sistem Günlüğü", font=("Roboto", 14, "bold")).pack(side="left")
        
        self.btn_toggle_errors = ctk.CTkSwitch(log_header, text="Sadece Hatalar", command=self.toggle_error_filter, 
                                              progress_color=THEME["danger"])
        self.btn_toggle_errors.pack(side="right")
        
        self.log_box = ctk.CTkTextbox(log_card, font=("Consolas", 12), state="disabled",
                                     fg_color="#0f172a", text_color="#e2e8f0", corner_radius=10)
        self.log_box.pack(fill="both", expand=True, padx=20, pady=(0, 20))

        # Progress bar artık progress_card içinde

    def _stat_card(self, parent, title, value, color, grid_col):
        card = ctk.CTkFrame(parent, fg_color=THEME["bg_card"], corner_radius=15, border_width=0)
        card.grid(row=0, column=grid_col, padx=5, sticky="ew")
        
        # Sol kenar çizgisi
        bar = ctk.CTkFrame(card, width=6, fg_color=color, corner_radius=10)
        bar.pack(side="left", fill="y", pady=10, padx=(10, 5))
        
        content = ctk.CTkFrame(card, fg_color="transparent")
        content.pack(side="left", fill="both", expand=True, pady=10, padx=5)
        
        ctk.CTkLabel(content, text=title, font=("Roboto", 12), text_color=THEME["text_muted"]).pack(anchor="w")
        lbl_val = ctk.CTkLabel(content, text=str(value), font=("Roboto", 24, "bold"), text_color=THEME["text_main"])
        lbl_val.pack(anchor="w")
        return lbl_val

    # ------------------------------
    # AYARLAR TABI (YENİ TASARIM)
    # ------------------------------
    def setup_settings(self):
        # Ana container - iki kolonlu grid (Sol: 40%, Sağ: 60%)
        self.tab_settings.grid_columnconfigure(0, weight=2)
        self.tab_settings.grid_columnconfigure(1, weight=3)
        self.tab_settings.grid_rowconfigure(0, weight=1)
        
        # === SOL PANEL: PARAMETRE EŞLEŞTİRME ===
        left_panel = ctk.CTkFrame(self.tab_settings, fg_color=THEME["bg_card"], corner_radius=20)
        left_panel.grid(row=0, column=0, padx=(20, 10), pady=20, sticky="nsew")
        left_panel.grid_columnconfigure(0, weight=1)
        left_panel.grid_rowconfigure(1, weight=1)
        
        # Başlık ve Ekle Butonu
        header_frame = ctk.CTkFrame(left_panel, fg_color="transparent")
        header_frame.grid(row=0, column=0, padx=25, pady=25, sticky="ew")
        header_frame.grid_columnconfigure(0, weight=1)
        
        title_lbl = ctk.CTkLabel(header_frame, text="Parametre Eşleştirme", 
                                font=("Roboto", 22, "bold"), text_color=THEME["text_main"])
        title_lbl.grid(row=0, column=0, sticky="w")
        
        add_btn = ctk.CTkButton(header_frame, text="+ Yeni Ekle", 
                               command=lambda: self.add_param_row("", ""),
                               fg_color=THEME["primary"], hover_color=THEME["primary_hover"],
                               width=120, height=35, font=("Roboto", 13, "bold"))
        add_btn.grid(row=0, column=1, sticky="e")
        
        # Scrollable Parametre Listesi
        self.scroll_params = ctk.CTkScrollableFrame(left_panel, fg_color="transparent", 
                                                   corner_radius=0)
        self.scroll_params.grid(row=1, column=0, padx=20, pady=(0, 20), sticky="nsew")
        
        # Varsayılan Parametreler (Thickness için ek boş, diğerleri standart)
        self.default_params = [
            ("", "B"),      # Thickness - ek boş
            ("H", "C"),    # H
            ("P1", "K"),   # P1
            ("D1", "L"),   # D1
            ("P2", "M"),   # P2
            ("D2", "N")    # D2
        ]
        for name, col in self.default_params:
            self.add_param_row(name, col)
        
        # Alt Araç Çubuğu
        toolbar = ctk.CTkFrame(left_panel, fg_color=THEME["bg_dark"], corner_radius=12, height=50)
        toolbar.grid(row=2, column=0, padx=20, pady=(0, 20), sticky="ew")
        
        clear_btn = ctk.CTkButton(toolbar, text="🧹 Tümünü Temizle", 
                                 command=self.clear_param_rows,
                                 fg_color="transparent", 
                                 text_color=THEME["danger"],
                                 hover_color=THEME["bg_card_hover"],
                                 height=35, font=("Roboto", 12))
        clear_btn.pack(side="left", padx=15, pady=7)
        
        # === SAĞ PANEL: CANLI ÖNİZLEME ===
        right_panel = ctk.CTkFrame(self.tab_settings, fg_color=THEME["bg_card"], corner_radius=20)
        right_panel.grid(row=0, column=1, padx=(10, 20), pady=20, sticky="nsew")
        right_panel.grid_columnconfigure(0, weight=1)
        right_panel.grid_rowconfigure(2, weight=1)
        
        # Başlık
        preview_title = ctk.CTkLabel(right_panel, text="Canlı Önizleme", 
                                    font=("Roboto", 22, "bold"), text_color=THEME["text_main"])
        preview_title.grid(row=0, column=0, padx=25, pady=(25, 15), sticky="w")
        
        # Sayfa Seçimi
        sheet_frame = ctk.CTkFrame(right_panel, fg_color="transparent")
        sheet_frame.grid(row=1, column=0, padx=25, pady=(0, 20), sticky="ew")
        
        sheet_label = ctk.CTkLabel(sheet_frame, text="Çalışma Sayfası:", 
                                  font=("Roboto", 13), text_color=THEME["text_muted"])
        sheet_label.pack(anchor="w", pady=(0, 8))
        
        self.combo_sheet = ctk.CTkComboBox(sheet_frame, values=["Dosya Bekleniyor..."], 
                                          command=self.on_sheet_change,
                                          fg_color=THEME["bg_dark"], 
                                          border_color=THEME["border"],
                                          button_color=THEME["primary"],
                                          button_hover_color=THEME["primary_hover"],
                                          height=40, font=("Roboto", 13))
        self.combo_sheet.pack(fill="x")
        
        # Veri Tablosu
        self.preview_box = ctk.CTkScrollableFrame(right_panel, fg_color=THEME["bg_dark"], 
                                                  corner_radius=15)
        self.preview_box.grid(row=2, column=0, padx=25, pady=(0, 25), sticky="nsew")
        
        # Empty State
        self.empty_state_frame = ctk.CTkFrame(self.preview_box, fg_color="transparent")
        self.empty_state_frame.pack(expand=True, fill="both", pady=80)
        
        empty_icon = ctk.CTkLabel(self.empty_state_frame, text="📊", font=("Arial", 48))
        empty_icon.pack(pady=(0, 15))
        
        empty_text = ctk.CTkLabel(self.empty_state_frame, 
                                 text="Veri önizlemesi için\nbir Excel dosyası seçin",
                                 font=("Roboto", 14), 
                                 text_color=THEME["text_muted"],
                                 justify="center")
        empty_text.pack()

    def add_param_row(self, default_name, default_col):
        """
        Daha şık, 'Kart' görünümlü parametre satırı.
        """
        # Kart Konteyner
        card = ctk.CTkFrame(self.scroll_params, fg_color=THEME["bg_card"], corner_radius=8)
        card.pack(fill="x", pady=4, padx=2)
        
        # İçerik için bir frame (pack kullanarak)
        content_frame = ctk.CTkFrame(card, fg_color="transparent")
        content_frame.pack(fill="x", padx=10, pady=10)
        
        # 1. Sütun Harfi (Badge Style)
        col_container = ctk.CTkFrame(content_frame, fg_color="transparent")
        col_container.pack(side="left", padx=(0, 8))
        
        ctk.CTkLabel(col_container, text="Sütun", font=("Arial", 9), text_color="gray").pack(anchor="w", pady=(0,2))
        col_entry = ctk.CTkEntry(col_container, width=45, height=32, 
                                 font=("Roboto", 13, "bold"), justify="center",
                                 fg_color=THEME["bg_dark"], 
                                 border_width=1, border_color=THEME["border"],
                                 corner_radius=6)
        col_entry.insert(0, default_col)
        col_entry.pack()
        
        # Ok İkonu
        arrow_lbl = ctk.CTkLabel(content_frame, text="➜", text_color="#888", font=("Arial", 16))
        arrow_lbl.pack(side="left", padx=8)
        
        # 2. Parametre Adı (Suffix)
        name_container = ctk.CTkFrame(content_frame, fg_color="transparent")
        name_container.pack(side="left", fill="x", expand=True, padx=(0, 8))
        
        ctk.CTkLabel(name_container, text="CATIA Parametre Adı (Ek)", font=("Arial", 9), text_color="gray").pack(anchor="w", pady=(0,2))
        name_entry = ctk.CTkEntry(name_container, height=32, font=("Roboto", 12),
                                  placeholder_text="Örn: Thickness", 
                                  fg_color=THEME["bg_dark"],
                                  border_width=1, border_color=THEME["border"],
                                  corner_radius=6)
        name_entry.insert(0, default_name)
        name_entry.pack(fill="x")
        
        # 3. Canlı ID Önizlemesi (Küçük bilgi)
        info_frame = ctk.CTkFrame(content_frame, fg_color="transparent")
        info_frame.pack(side="left", padx=(0, 5))
        ctk.CTkLabel(info_frame, text="→", font=("Arial", 12), text_color="#666").pack()
        info_lbl = ctk.CTkLabel(info_frame, text=f"ID{default_name}", 
                               font=("Consolas", 10, "bold"), 
                               text_color=THEME["primary"])
        info_lbl.pack()
        
        # 4. Silme Butonu (Ghost Style)
        del_btn = ctk.CTkButton(content_frame, text="✕", width=32, height=32,
                                fg_color="transparent", text_color=THEME["text_muted"],
                                hover_color=THEME["danger"],
                                font=("Arial", 16, "bold"),
                                corner_radius=6,
                                command=lambda: self.delete_param_row(card))
        del_btn.pack(side="right", padx=(5, 0))
        
        # Referansları Sakla
        def normalize_col(event=None):
            raw = col_entry.get().upper()
            normalized = "".join([ch for ch in raw if ch.isalpha()])
            if raw != normalized:
                col_entry.delete(0, "end")
                col_entry.insert(0, normalized)
            if normalized:
                col_entry.configure(text_color=THEME["text_main"])
            else:
                col_entry.configure(text_color=THEME["danger"])

        def update_info(event=None):
            suf = name_entry.get().strip()
            info_lbl.configure(text=f"ID{suf}" if suf else "ID...")

        col_entry.bind("<KeyRelease>", normalize_col)
        name_entry.bind("<KeyRelease>", update_info)
        normalize_col()
        update_info()

        self.param_rows.append({
            "frame": card,
            "col": col_entry,
            "name": name_entry,
            "info": info_lbl,
            "delete_btn": del_btn
        })

    def delete_param_row(self, frame):
        frame.destroy()
        self.param_rows = [r for r in self.param_rows if r["frame"] != frame]

    def clear_param_rows(self):
        for row in self.param_rows:
            row["frame"].destroy()
        self.param_rows = []
        for name, col in self.default_params:
            self.add_param_row(name, col)

    def set_controls_state(self, enabled: bool):
        state = "normal" if enabled else "disabled"
        combo_state = "normal" if enabled else "disabled"
        if hasattr(self, "btn_file"):
            self.btn_file.configure(state=state)
        if hasattr(self, "btn_add_param"):
            self.btn_add_param.configure(state=state)
        if hasattr(self, "btn_sort_params"):
            self.btn_sort_params.configure(state=state)
        if hasattr(self, "btn_clear_params"):
            self.btn_clear_params.configure(state=state)
        if hasattr(self, "combo_sheet"):
            self.combo_sheet.configure(state=combo_state)
        for row in self.param_rows:
            row["col"].configure(state=state)
            row["name"].configure(state=state)
            row["delete_btn"].configure(state=state)

    def set_running_state(self, running: bool):
        # Yeni tasarımda status badge yok, sadece buton durumlarını güncelle
        pass

    def update_summary_label(self):
        # Yeni tasarımda summary label yok, gerekirse eklenebilir
        pass

    def on_sheet_change(self, value: str):
        self.config["sheet_name"] = value
        self.update_summary_label()

    def update_error_button_text(self):
        # Yeni tasarımda switch kullanıyoruz, text güncellemesi gerekmiyor
        pass

    # ------------------------------
    # İŞLEVLER
    # ------------------------------
    def select_file(self):
        was_topmost = bool(self.attributes("-topmost"))
        self.attributes("-topmost", False)
        path = ctk.filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx *.xlsm *.xls")])
        
        if path:
            if self.validate_excel_file(path):
                self.load_excel_file(path)
        
        self.attributes("-topmost", was_topmost)
    
    def load_excel_file(self, path):
        """Dosya yükleme işlemini gerçekleştirir"""
        self.selected_file = path
        if hasattr(self, 'lbl_file_info'):
            self.lbl_file_info.configure(text=f"✅ {os.path.basename(path)}")
        if hasattr(self, 'btn_run'):
            self.btn_run.configure(state="normal")
        if hasattr(self, 'btn_remove_file'):
            self.btn_remove_file.configure(state="normal")
        self.log(f"Dosya seçildi: {os.path.basename(path)}", "info")
        self.show_toast("Dosya Yüklendi", f"{os.path.basename(path)} hazır.", type="success")
        
        # Tabı Ayarlar'a geçir (opsiyonel)
        # self.tab_view.set("  ⚙️ Ayarlar  ")
        self.log("Önizleme oluşturuluyor...", "info")
        ExcelPreviewLoader(self, path).start()
        self.update_summary_label()
    
    def remove_file(self):
        """Seçili dosyayı kaldır"""
        if not self.selected_file:
            return
        
        # Onay al
        if messagebox.askyesno("Dosyayı Kaldır", 
                              f"'{os.path.basename(self.selected_file)}' dosyasını kaldırmak istediğinize emin misiniz?"):
            # Dosyayı temizle
            self.selected_file = None
            
            # UI'yi güncelle
            if hasattr(self, 'lbl_file_info'):
                self.lbl_file_info.configure(text="Dosya seçilmedi")
            if hasattr(self, 'btn_run'):
                self.btn_run.configure(state="disabled")
            if hasattr(self, 'btn_remove_file'):
                self.btn_remove_file.configure(state="disabled")
            
            # Önizlemeyi tamamen temizle
            if hasattr(self, 'preview_box'):
                # preview_box içindeki TÜM widget'ları temizle
                for widget in self.preview_box.winfo_children():
                    try:
                        widget.destroy()
                    except:
                        pass
                
                # Referansları temizle
                self.preview_table = None
                self.empty_state_frame = None
                
                # Yeni empty state oluştur
                self.empty_state_frame = ctk.CTkFrame(self.preview_box, fg_color="transparent")
                self.empty_state_frame.pack(expand=True, fill="both", pady=80)
                
                empty_icon = ctk.CTkLabel(self.empty_state_frame, text="📊", font=("Arial", 48))
                empty_icon.pack(pady=(0, 15))
                
                empty_text = ctk.CTkLabel(self.empty_state_frame, 
                                         text="Veri önizlemesi için\nbir Excel dosyası seçin",
                                         font=("Roboto", 14), 
                                         text_color=THEME["text_muted"],
                                         justify="center")
                empty_text.pack()
            
            # Sayfa listesini temizle
            if hasattr(self, 'combo_sheet'):
                self.combo_sheet.configure(values=["Dosya Bekleniyor..."])
                self.combo_sheet.set("Dosya Bekleniyor...")
            
            self.log("Dosya kaldırıldı", "info")
            self.show_toast("Dosya Kaldırıldı", "Dosya seçimi temizlendi.", type="info")
    
    def validate_excel_file(self, path):
        """Excel dosyasını doğrular"""
        try:
            # Dosya varlık kontrolü
            if not os.path.exists(path):
                self.log("Dosya bulunamadı!", "error")
                APP_LOGGER.error(f"Dosya bulunamadı: {path}")
                return False
            
            # Format kontrolü
            if not path.lower().endswith(('.xlsx', '.xlsm', '.xls')):
                self.log("Geçersiz dosya formatı! (.xlsx, .xlsm veya .xls olmalı)", "error")
                APP_LOGGER.error(f"Geçersiz format: {path}")
                return False
            
            # Dosya boyutu kontrolü (max 100MB)
            file_size = os.path.getsize(path)
            max_size = 100 * 1024 * 1024  # 100 MB
            if file_size > max_size:
                size_mb = file_size / (1024 * 1024)
                self.log(f"Dosya çok büyük! ({size_mb:.1f}MB - Max: 100MB)", "error")
                APP_LOGGER.error(f"Dosya çok büyük: {size_mb:.1f}MB")
                return False
            
            # Dosya okuma izni kontrolü
            if not os.access(path, os.R_OK):
                self.log("Dosya okunamıyor! İzin hatası.", "error")
                APP_LOGGER.error(f"Okuma izni yok: {path}")
                return False
            
            APP_LOGGER.info(f"Dosya validasyonu başarılı: {path} ({file_size / 1024:.1f}KB)")
            return True
            
        except Exception as e:
            self.log(f"Dosya validasyon hatası: {e}", "error")
            APP_LOGGER.error(f"Validasyon hatası: {e}\n{traceback.format_exc()}")
            return False

    def update_ui_with_excel_data(self, sheets, data):
        # Sayfa listesini güncelle
        self.combo_sheet.configure(values=sheets)
        if sheets:
            self.combo_sheet.set(sheets[0])
            self.on_sheet_change(self.combo_sheet.get())

        # preview_box içindeki TÜM widget'ları temizle
        if hasattr(self, 'preview_box'):
            for widget in self.preview_box.winfo_children():
                try:
                    widget.destroy()
                except:
                    pass
        
        # Referansları temizle
        self.preview_table = None
        self.empty_state_frame = None
            
        if not data:
            self.preview_table = ctk.CTkFrame(self.preview_box, fg_color="transparent")
            self.preview_table.pack(fill="both", expand=True, padx=15, pady=15)
            ctk.CTkLabel(self.preview_table, text="Veri okunamadı.", text_color="red").pack()
            return
        
        # Daha iyi tablo görünümü için container
        table_wrapper = ctk.CTkFrame(self.preview_box, fg_color="transparent")
        table_wrapper.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Canvas ve scrollbar container
        canvas_container = ctk.CTkFrame(table_wrapper, fg_color="transparent")
        canvas_container.pack(fill="both", expand=True)
        
        # Canvas için frame
        canvas_frame = ctk.CTkFrame(canvas_container, fg_color=THEME["bg_dark"], corner_radius=10)
        canvas_frame.pack(fill="both", expand=True)
        
        # Canvas ve scrollbar
        canvas = Canvas(canvas_frame, bg=THEME["bg_dark"], highlightthickness=0, 
                       borderwidth=0, relief="flat")
        h_scrollbar = ctk.CTkScrollbar(canvas_frame, orientation="horizontal", command=canvas.xview)
        v_scrollbar = ctk.CTkScrollbar(canvas_frame, orientation="vertical", command=canvas.yview)
        
        scrollable_frame = ctk.CTkFrame(canvas, fg_color="transparent")
        
        canvas_window = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        
        def configure_scroll_region(event=None):
            canvas.update_idletasks()
            bbox = canvas.bbox("all")
            if bbox:
                canvas.configure(scrollregion=bbox)
        
        def configure_canvas_width(event):
            # Canvas window genişliğini içeriğin genişliğine göre ayarla (yatay scroll için)
            scrollable_frame.update_idletasks()
            frame_width = scrollable_frame.winfo_reqwidth()
            # İçeriğin genişliğine göre canvas window'u ayarla
            if frame_width > 0:
                canvas.itemconfig(canvas_window, width=frame_width)
            configure_scroll_region()
        
        scrollable_frame.bind("<Configure>", configure_scroll_region)
        canvas.bind("<Configure>", configure_canvas_width)
        canvas.configure(xscrollcommand=h_scrollbar.set, yscrollcommand=v_scrollbar.set)
        
        # Layout
        canvas.grid(row=0, column=0, sticky="nsew")
        v_scrollbar.grid(row=0, column=1, sticky="ns")
        h_scrollbar.grid(row=1, column=0, sticky="ew")
        canvas_frame.grid_columnconfigure(0, weight=1)
        canvas_frame.grid_rowconfigure(0, weight=1)
        
        # Tablo
        self.preview_table = ctk.CTkFrame(scrollable_frame, fg_color="transparent")
        self.preview_table.pack(fill="both", expand=True, padx=5, pady=5)
        
        num_cols = len(data[0]) if data else 0
        
        # Grid column ayarları
        self.preview_table.grid_columnconfigure(0, minsize=50)  # Satır numarası
        for c_idx in range(num_cols):
            self.preview_table.grid_columnconfigure(c_idx + 1, minsize=80)  # Her sütun
        
        # Başlık satırı
        row_header = ctk.CTkLabel(self.preview_table, text="#", 
                                 fg_color=THEME["bg_card_hover"],
                                 font=("Roboto", 10, "bold"), height=32, corner_radius=6,
                                 width=50)
        row_header.grid(row=0, column=0, padx=3, pady=3, sticky="ew")
        
        # Sütun başlıkları
        for c_idx in range(num_cols):
            col_letter = num2col(c_idx + 1)
            header = ctk.CTkLabel(self.preview_table, text=col_letter, 
                                  fg_color=THEME["bg_card_hover"], 
                                  font=("Roboto", 10, "bold"), height=32, 
                                  corner_radius=6, width=80)
            header.grid(row=0, column=c_idx+1, padx=3, pady=3, sticky="ew")

        # Veri satırları
        for r_idx, row_data in enumerate(data):
            bg_row = THEME["bg_card"] if r_idx % 2 == 0 else THEME["bg_card_hover"]
            
            # Satır numarası
            row_num = ctk.CTkLabel(self.preview_table, text=str(r_idx+1), 
                                   text_color=THEME["text_muted"], 
                                   font=("Roboto", 9),
                                   fg_color=bg_row, height=28, width=50)
            row_num.grid(row=r_idx+1, column=0, padx=3, pady=2, sticky="ew")
            
            # Hücreler
            for c_idx, cell_val in enumerate(row_data):
                val_str = str(cell_val) if cell_val is not None else ""
                # Metin kısaltma
                if len(val_str) > 15: 
                    val_str = val_str[:13] + "..."
                
                cell = ctk.CTkLabel(self.preview_table, text=val_str, height=28,
                                    fg_color=bg_row, font=("Consolas", 10), 
                                    width=80, anchor="w", padx=6)
                cell.grid(row=r_idx+1, column=c_idx+1, padx=3, pady=2, sticky="ew")
        
        # Canvas scroll ayarı - düzgün çalışması için
        def update_scroll_region(event=None):
            canvas.update_idletasks()
            bbox = canvas.bbox("all")
            if bbox:
                canvas.configure(scrollregion=bbox)
        
        # Scrollable frame boyutunu ayarla
        def on_frame_configure(event):
            # İçeriğin genişliğine göre canvas window'u ayarla
            scrollable_frame.update_idletasks()
            frame_width = scrollable_frame.winfo_reqwidth()
            if frame_width > 0:
                canvas.itemconfig(canvas_window, width=frame_width)
            update_scroll_region()
        
        def on_canvas_configure(event):
            # Canvas genişliği değiştiğinde içeriğin genişliğine göre canvas window'u ayarla
            scrollable_frame.update_idletasks()
            frame_width = scrollable_frame.winfo_reqwidth()
            if frame_width > 0:
                canvas.itemconfig(canvas_window, width=frame_width)
            update_scroll_region()
        
        scrollable_frame.bind("<Configure>", on_frame_configure)
        canvas.bind("<Configure>", on_canvas_configure)
        
        # İlk güncelleme (biraz gecikme ile)
        def initial_update():
            canvas.update_idletasks()
            scrollable_frame.update_idletasks()
            update_scroll_region()
        
        canvas.after(100, initial_update)
        
        # Mouse wheel desteği (tüm platformlar için)
        def on_mousewheel(event):
            # Windows ve macOS
            if sys.platform == "win32":
                if event.state & 0x0001:  # Shift tuşu basılı
                    canvas.xview_scroll(int(-1 * (event.delta / 120)), "units")
                else:
                    canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            elif sys.platform == "darwin":
                if event.state & 0x0001:  # Shift tuşu basılı
                    canvas.xview_scroll(int(-1 * (event.delta / 120)), "units")
                else:
                    canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            else:
                # Linux
                if event.num == 4:
                    canvas.yview_scroll(-1, "units")
                elif event.num == 5:
                    canvas.yview_scroll(1, "units")
        
        if sys.platform == "win32" or sys.platform == "darwin":
            canvas.bind("<MouseWheel>", on_mousewheel)
        else:
            canvas.bind("<Button-4>", on_mousewheel)
            canvas.bind("<Button-5>", on_mousewheel)
        
        # Horizontal scroll için Shift+Wheel
        def on_shift_wheel(event):
            canvas.xview_scroll(int(-1 * (event.delta / 120)), "units")
        
        if sys.platform in ["win32", "darwin"]:
            canvas.bind("<Shift-MouseWheel>", on_shift_wheel)
        self.log("Önizleme yüklendi.", "success")

    def start_process(self):
        """İşlemi başlat - Gelişmiş validasyon ile"""
        try:
            # CATIA kontrolü (Test modu değilse)
            if not TEST_MODE:
                catia_available = False
                try:
                    if WIN32COM_AVAILABLE:
                        import win32com.client
                        catia = win32com.client.GetActiveObject("CATIA.Application")
                        if catia:
                            # ActiveDocument kontrolü
                            try:
                                doc = catia.ActiveDocument
                                if doc:
                                    catia_available = True
                            except:
                                pass
                except Exception as e:
                    APP_LOGGER.warning(f"CATIA kontrolü: {e}")
                
                if not catia_available:
                    self.show_toast("Hata", "CATIA açık değil!\n\nLütfen önce CATIA'yı açın ve bir döküman açın.", type="error")
                    self.log("CATIA açık değil veya döküman yok!", "error")
                    return
            
            # Parametreleri topla
            dynamic_params = []
            for row in self.param_rows:
                suffix = row["name"].get().strip()
                col = row["col"].get().strip()
                if col:
                    # Sütun validasyonu
                    if not col.isalpha():
                        self.show_toast("Hata", f"Geçersiz sütun: '{col}'", type="error")
                        return
                    dynamic_params.append((suffix, col))
            
            # Temel validasyonlar
            if not self.selected_file:
                self.show_toast("Uyarı", "Lütfen önce bir Excel dosyası seçin!", type="warning")
                return
            
            if not dynamic_params:
                self.show_toast("Uyarı", "Ayarlar sekmesinden parametre eşleştirmesi yapın!", type="warning")
                return
            
            # Dosya tekrar validate et
            if not self.validate_excel_file(self.selected_file):
                self.show_toast("Hata", "Dosya validasyonu başarısız!", type="error")
                return
            
            self.config["sheet_name"] = self.combo_sheet.get()
            
            if not self.config["sheet_name"]:
                self.show_toast("Uyarı", "Lütfen bir Excel sayfası seçin!", type="warning")
                return
            
            # UI'yi hazırla
            self.tab_view.set("  🚀 Monitör  ")
            self.set_controls_state(False)
            self.set_running_state(True)
            self.btn_run.configure(state="disabled", text="ÇALIŞIYOR...", fg_color=THEME["bg_card_hover"])
            self.log_entries.clear()
            self.render_log()
            self.current_run_has_error = False
            self.btn_stop.configure(state="normal")
            
            # İşlemi başlat
            self.start_time = time.time()
            APP_LOGGER.info(f"İşlem başlatılıyor - Dosya: {self.selected_file}, Sayfa: {self.config['sheet_name']}")
            
            self.worker = WorkerThread(self, self.selected_file, self.config, dynamic_params)
            self.worker.start()
            
        except Exception as e:
            APP_LOGGER.error(f"İşlem başlatma hatası: {e}\n{traceback.format_exc()}")
            self.show_toast("Hata", f"İşlem başlatılamadı:\n{str(e)}", type="error")

    # (Diğer yardımcı fonksiyonlar: log, update_stats vs. aynı kalır)
    def log(self, msg, type="info"):
        icons = {"info": "ℹ", "update": "⚡", "error": "✖", "success": "✔"}
        ts = datetime.datetime.now().strftime("%H:%M:%S")
        self.log_entries.append({"ts": ts, "type": type, "icon": icons.get(type, ""), "msg": msg})
        
        max_logs = 5000
        if len(self.log_entries) > max_logs:
            self.log_entries = self.log_entries[-max_logs:]
        
        if type == "error":
            self.current_run_has_error = True
            # Btn rengini değiştirme, sadece status badge veya log yeterli
        
        if not hasattr(self, '_log_render_count'):
            self._log_render_count = 0
        self._log_render_count += 1
        if self._log_render_count % 10 == 0 or type in ["error", "success"]:
            self.render_log()

    def render_log(self):
        """Log render'ını optimize et - sadece son 1000 satırı göster"""
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        
        filtered = self.log_entries
        if self.show_errors_only:
            filtered = [e for e in self.log_entries if e["type"] == "error"]
        
        max_display = 1000
        if len(filtered) > max_display:
            filtered = filtered[-max_display:]
            self.log_box.insert("end", f"... ({len(self.log_entries) - max_display} eski log gizlendi) ...\n")
        
        log_text = "\n".join([f"[{e['ts']}] {e['icon']} {e['msg']}" for e in filtered])
        self.log_box.insert("end", log_text + "\n")
        self.log_box.see("end")
        self.log_box.configure(state="disabled")

    def toggle_error_filter(self):
        self.show_errors_only = not self.show_errors_only
        self.render_log()
        
    # update_stats metodunu yukarıda tanımlamıştık, burada tekrar etmeye gerek yok
    # update_max_progress vb. metodlar...

    def update_max_progress(self, val): 
        self.total_work = val
        # start_time sadece start_process'te ayarlanmalı, burada değil
        if hasattr(self, 'progress_bar'):
            self.progress_bar.set(0)
        if hasattr(self, 'lbl_progress'):
            self.lbl_progress.configure(text=f"0 / {val}")
    
    def update_stats(self, current, updates, errors):
        """İstatistikleri ve progress bar'ı güncelle"""
        # Progress bar
        if hasattr(self, 'progress_bar') and self.total_work > 0:
            progress = current / self.total_work
            self.progress_bar.set(progress)
        if hasattr(self, 'lbl_progress'):
            self.lbl_progress.configure(text=f"{current} / {self.total_work}")
        
        # İstatistik kartları
        if hasattr(self, 'card_success'):
            self.card_success.configure(text=str(updates))
        if hasattr(self, 'card_error'):
            self.card_error.configure(text=str(errors))
        if hasattr(self, 'card_time'):
            # Geçen süreyi hesapla
            if hasattr(self, 'start_time') and self.start_time > 0:
                elapsed = max(0, time.time() - self.start_time)
                minutes = int(elapsed // 60)
                seconds = int(elapsed % 60)
                time_str = f"{minutes:02d}:{seconds:02d}"
            else:
                time_str = "00:00"
            self.card_time.configure(text=time_str)
        
    def finish_process(self):
        self.set_controls_state(True)
        self.btn_stop.configure(state="disabled")
        self.set_running_state(False)
        
        fg = THEME["danger"] if self.current_run_has_error else THEME["success"]
        self.btn_run.configure(state="normal", text="🚀 BAŞLAT", fg_color=fg)
        
        elapsed = time.time() - self.start_time
        total_updates = int(self.card_success.cget("text") or 0)
        total_errors = int(self.card_error.cget("text") or 0)
        
        self.log("İşlem tamamlandı.", "success")
        
        try:
            self.auto_save_results(total_updates, total_errors, elapsed)
        except Exception as e:
            APP_LOGGER.error(f"Otomatik kayıt hatası: {e}")
        
        if total_errors > 0:
            self.show_toast("İşlem Tamamlandı", f"{total_errors} hata oluştu. Logları kontrol edin.", type="error")
        else:
            self.show_toast("İşlem Başarılı", "Tüm veriler aktarıldı.", type="success")
    
    def auto_save_results(self, updates, errors, elapsed_time):
        """İşlem sonuçlarını otomatik kaydeder"""
        if not os.path.exists("Results"):
            os.makedirs("Results")
        
        filename = os.path.join("Results", f"result_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.txt")
        try:
            with open(filename, "w", encoding="utf-8") as f:
                f.write(f"AFT Sizing Automation - İşlem Sonuçları\n")
                f.write(f"{'='*50}\n\n")
                f.write(f"Tarih: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Dosya: {os.path.basename(self.selected_file) if self.selected_file else 'N/A'}\n")
                f.write(f"Sayfa: {self.config.get('sheet_name', 'N/A')}\n\n")
                f.write(f"Toplam İşlenen: {self.total_work}\n")
                f.write(f"Başarılı: {updates}\n")
                f.write(f"Hatalar: {errors}\n")
                f.write(f"Başarı Oranı: {(updates/self.total_work*100) if self.total_work > 0 else 0:.1f}%\n")
                f.write(f"Geçen Süre: {time.strftime('%M:%S', time.gmtime(elapsed_time))}\n")
            self.log(f"Sonuçlar otomatik kaydedildi: {os.path.basename(filename)}", "info")
        except Exception as e:
            self.log(f"Sonuç kaydedilemedi: {e}", "error")
    
    def stop_process(self):
        if self.worker:
            self.worker.stop()
            self.btn_stop.configure(state="disabled")
            self.log("Durdurma talebi gönderildi.", "info")
    
    
    def on_closing(self):
        """Uygulama kapatılırken"""
        APP_LOGGER.info("Uygulama kapatılıyor...")
        
        # Çalışan işlem varsa sor
        if self.worker and self.worker.is_alive():
            if messagebox.askyesnocancel("Çıkış", "İşlem devam ediyor!\n\nYine de çıkmak istiyor musunuz?"):
                self.worker.stop()
                self.worker.join(timeout=2)
                self.destroy()
        else:
            self.destroy()

if __name__ == "__main__":
    APP_LOGGER.info("=" * 50)
    APP_LOGGER.info("AFT Sizing Automation - Başlatılıyor")
    APP_LOGGER.info(f"Test Modu: {TEST_MODE}")
    APP_LOGGER.info(f"openpyxl: {'Yüklü' if OPENPYXL_AVAILABLE else 'Yüklü Değil'}")
    APP_LOGGER.info(f"win32com: {'Yüklü' if WIN32COM_AVAILABLE else 'Yüklü Değil'}")
    APP_LOGGER.info("=" * 50)
    
    try:
        app = AutomationSuite()
        # Pencereyi ekranın önüne getir
        app.lift()
        app.attributes("-topmost", True)
        app.after(100, lambda: app.attributes("-topmost", False))  # 100ms sonra topmost'u kapat
        app.focus_force()
        app.mainloop()
    except Exception as e:
        APP_LOGGER.critical(f"Kritik uygulama hatası: {e}\n{traceback.format_exc()}")
        messagebox.showerror("Kritik Hata", f"Uygulama başlatılamadı:\n{str(e)}")
    finally:
        APP_LOGGER.info("Uygulama kapatıldı")