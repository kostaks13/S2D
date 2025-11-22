import customtkinter as ctk
import threading
import time
import datetime
import os
import string
import json
import logging
import traceback
from logging.handlers import RotatingFileHandler
from tkinter import filedialog, messagebox
from tkinterdnd2 import DND_FILES, TkinterDnD
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️ openpyxl bulunamadı. 'pip install openpyxl' ile yükleyin (daha hızlı Excel okuma için)")
try:
    import win32com.client
    import pythoncom
    WIN32COM_AVAILABLE = True
except ImportError:
    WIN32COM_AVAILABLE = False
    print("⚠️ win32com bulunamadı. Test modu dışında çalışmayabilir.")

# ==========================================
# AYARLAR & YARDIMCILAR
# ==========================================
TEST_MODE = True  # True: Test Modu (Excel gerekmez) / False: Gerçek Mod

# ==========================================
# LOGGING SİSTEMİ
# ==========================================
def setup_logger():
    """Gelişmiş logging sistemi - Dosya ve konsol çıktısı"""
    if not os.path.exists("Logs"):
        os.makedirs("Logs")
    
    logger = logging.getLogger("CATIA_Automation")
    logger.setLevel(logging.DEBUG)
    
    # Eğer logger zaten handler'lara sahipse, tekrar ekleme
    if logger.handlers:
        return logger
    
    # Dosya handler (Rotating - Max 5MB, 5 yedek)
    log_file = os.path.join("Logs", "catia_automation.log")
    fh = RotatingFileHandler(log_file, maxBytes=5*1024*1024, backupCount=5, encoding='utf-8')
    fh.setLevel(logging.DEBUG)
    
    # Formatter
    formatter = logging.Formatter(
        '%(asctime)s - %(name)s - [%(levelname)s] - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    fh.setFormatter(formatter)
    
    logger.addHandler(fh)
    return logger

# Global logger
APP_LOGGER = setup_logger()

# ==========================================
# VERİ DOĞRULAMA
# ==========================================
def validate_parameter_value(value, param_name):
    """Parametre değerini CATIA'ya göndermeden önce doğrula"""
    try:
        float_val = float(value)
        
        # Özel doğrulama kuralları
        if "Thickness" in param_name or param_name.startswith("T"):
            if float_val <= 0:
                raise ValueError(f"{param_name} pozitif olmalı")
        
        if "Angle" in param_name:
            if not -360 <= float_val <= 360:
                raise ValueError(f"{param_name} -360 ile 360 arasında olmalı")
        
        return float_val
    except ValueError as e:
        raise ValueError(f"{param_name} için geçersiz değer '{value}': {str(e)}")

# ==========================================
# EXCEL İŞLEMLERİ
# ==========================================
def read_excel_openpyxl(file_path, sheet_name=None):
    """openpyxl ile hızlı Excel okuma"""
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        # Tüm satırları oku (ilk satır header)
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:  # İlk sütun (ID) boş değilse
                data.append(row)
        
        # Sayfa isimlerini al
        sheets = wb.sheetnames
        wb.close()
        
        return data, sheets
    except Exception as e:
        APP_LOGGER.error(f"openpyxl okuma hatası: {e}")
        raise

def read_excel_preview_openpyxl(file_path, max_rows=10):
    """openpyxl ile önizleme okuma (ilk N satır)"""
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        
        data = []
        for idx, row in enumerate(ws.iter_rows(values_only=True)):
            if idx >= max_rows:
                break
            data.append(list(row) if row else [])
        
        sheets = wb.sheetnames
        wb.close()
        
        return data, sheets
    except Exception as e:
        APP_LOGGER.error(f"openpyxl önizleme hatası: {e}")
        raise

def generate_excel_template(params, output_path):
    """Kullanıcı için Excel şablonu oluştur"""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl yüklü değil. 'pip install openpyxl' ile yükleyin.")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Visualization Data"
    
    # Header satırı
    headers = ["ID"] + [param[0] for param in params]
    ws.append(headers)
    
    # Stil ekle (header'ı kalın yap)
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3B8ED0", end_color="3B8ED0", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    # Örnek satırlar (10 adet)
    for i in range(1, 11):
        row_data = [f"Rib_{i}"] + ["0.0"] * len(params)
        ws.append(row_data)
    
    # Sütun genişliklerini ayarla
    ws.column_dimensions['A'].width = 15
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12
    
    wb.save(output_path)
    APP_LOGGER.info(f"Excel şablonu oluşturuldu: {output_path}")

# ==========================================
# RAPOR OLUŞTURMA
# ==========================================
def generate_html_report(log_entries, stats, output_path):
    """HTML formatında rapor oluştur"""
    html_content = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>CATIA Automation - İşlem Raporu</title>
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #1a1a1a 0%, #2d2d2d 100%);
            color: #ffffff;
            padding: 20px;
            margin: 0;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background: #2b2b2b;
            border-radius: 15px;
            padding: 30px;
            box-shadow: 0 10px 40px rgba(0,0,0,0.5);
        }}
        h1 {{
            color: #3B8ED0;
            border-bottom: 3px solid #3B8ED0;
            padding-bottom: 10px;
        }}
        .stats {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 15px;
            margin: 25px 0;
        }}
        .stat-card {{
            background: #333;
            padding: 20px;
            border-radius: 10px;
            text-align: center;
            border-left: 4px solid #3B8ED0;
        }}
        .stat-value {{
            font-size: 2em;
            font-weight: bold;
            color: #3B8ED0;
        }}
        .stat-label {{
            color: #aaa;
            font-size: 0.9em;
            margin-top: 5px;
        }}
        .log-section {{
            margin-top: 30px;
            background: #1e1e1e;
            padding: 20px;
            border-radius: 10px;
            max-height: 500px;
            overflow-y: auto;
        }}
        .log-entry {{
            padding: 8px 12px;
            margin: 5px 0;
            border-left: 3px solid #555;
            background: #252525;
            font-family: 'Consolas', monospace;
            font-size: 0.9em;
        }}
        .log-error {{
            border-left-color: #cf6679;
            background: #3d1f24;
        }}
        .log-success {{
            border-left-color: #2ECC71;
            background: #1f3d24;
        }}
        .log-update {{
            border-left-color: #F39C12;
            background: #3d3324;
        }}
        .timestamp {{
            color: #888;
            margin-right: 10px;
        }}
        .footer {{
            margin-top: 30px;
            text-align: center;
            color: #666;
            font-size: 0.85em;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>🚀 CATIA Automation Suite - İşlem Raporu</h1>
        
        <div class="stats">
            <div class="stat-card">
                <div class="stat-value">{stats.get('total', 0)}</div>
                <div class="stat-label">Toplam İşlenen</div>
            </div>
            <div class="stat-card">
                <div class="stat-value">{stats.get('updates', 0)}</div>
                <div class="stat-label">Başarılı</div>
            </div>
            <div class="stat-card">
                <div class="stat-value">{stats.get('errors', 0)}</div>
                <div class="stat-label">Hatalar</div>
            </div>
            <div class="stat-card">
                <div class="stat-value">{stats.get('success_rate', 0):.1f}%</div>
                <div class="stat-label">Başarı Oranı</div>
            </div>
            <div class="stat-card">
                <div class="stat-value">{stats.get('elapsed_time', '00:00')}</div>
                <div class="stat-label">Geçen Süre</div>
            </div>
        </div>
        
        <h2>📋 İşlem Detayları</h2>
        <p><strong>Dosya:</strong> {stats.get('filename', 'N/A')}</p>
        <p><strong>Sayfa:</strong> {stats.get('sheet', 'N/A')}</p>
        <p><strong>Tarih:</strong> {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        
        <h2>📜 Log Kayıtları</h2>
        <div class="log-section">
"""
    
    # Log girişlerini ekle
    for entry in log_entries:
        log_class = f"log-{entry['type']}" if entry['type'] in ['error', 'success', 'update'] else ""
        html_content += f"""
            <div class="log-entry {log_class}">
                <span class="timestamp">[{entry['ts']}]</span>
                <span>{entry['icon']} {entry['msg']}</span>
            </div>
"""
    
    html_content += """
        </div>
        
        <div class="footer">
            <p>CATIA Automation Suite v4.5 Pro - © 2025</p>
        </div>
    </div>
</body>
</html>
"""
    
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html_content)
    
    APP_LOGGER.info(f"HTML raporu oluşturuldu: {output_path}")

def col2num(col_str):
    """Harfi sayıya çevirir (A->1, Z->26, AA->27)"""
    num = 0
    col_str = col_str.strip().upper()
    for c in col_str:
        if c in string.ascii_letters:
            num = num * 26 + (ord(c) - ord('A')) + 1
    return num

def num2col(n):
    """Sayıyı harfe çevirir (1->A, 27->AA)"""
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string

# ==========================================
# ARKA PLAN İŞÇİSİ (DATA OKUMA & YAZMA)
# ==========================================
class WorkerThread(threading.Thread):
    def __init__(self, app, excel_path, config, dynamic_params):
        super().__init__()
        self.app = app
        self.excel_path = excel_path
        self.config = config
        self.dynamic_params = dynamic_params # List of tuples: (Suffix, ColChar)
        self.running = True
        self.daemon = True

    def run(self):
        if TEST_MODE:
            self.run_simulation()
        else:
            self.run_real_process()

    def run_simulation(self):
        sheet_name = self.config.get("sheet_name", "Sheet1")
        self.app.log(f"Başladı: {sheet_name}", "info")
        
        # Dinamik parametreleri logla
        msg = "Parametreler: "
        for suffix, col in self.dynamic_params:
            msg += f"[{suffix}->{col}] "
        self.app.log(msg, "info")
        
        time.sleep(1)
        total = 50
        self.app.after(0, self.app.update_max_progress, total)
        
        updates = 0
        errors = 0
        
        for i in range(1, total + 1):
            if not self.running: break
            time.sleep(0.05)
            self.app.after(0, self.app.update_stats, i, updates, errors)
            
            if i % 5 == 0:
                updates += 1
                # Örnek: İlk dinamik parametreyi update etmiş gibi yapalım
                p_name = f"Rib_{i}_{self.dynamic_params[0][0]}" if self.dynamic_params else "Param"
                self.app.after(0, self.app.log, f"{p_name} güncellendi.", "update")

        self.app.after(0, self.app.finish_process)

    def run_real_process(self):
        """Gerçek işlem - Excel okuma ve CATIA yazma"""
        excel = None
        catia = None
        
        try:
            APP_LOGGER.info("İşlem başlatıldı")
            
            # openpyxl ile okuma (varsa, daha hızlı)
            if OPENPYXL_AVAILABLE:
                APP_LOGGER.info("openpyxl ile Excel okunuyor...")
                self.run_with_openpyxl()
                return
            
            # Fallback: win32com ile okuma
            if not WIN32COM_AVAILABLE:
                raise ImportError("win32com bulunamadı. 'pip install pywin32' ile yükleyin.")
            
            APP_LOGGER.info("win32com ile Excel okunuyor...")
            import win32com.client
            import pythoncom
            pythoncom.CoInitialize()

            # Excel'i görünmez modda aç ve performans ayarları
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False  # Görünmez mod - daha hızlı
            excel.ScreenUpdating = False  # Ekran güncellemelerini kapat
            excel.DisplayAlerts = False  # Uyarıları kapat
            excel.EnableEvents = False  # Event'leri kapat
            excel.Calculation = -4135  # xlCalculationManual (otomatik hesaplamayı kapat)
            
            wb = excel.Workbooks.Open(self.excel_path, ReadOnly=True)  # Read-only aç - daha hızlı
            target_sheet = self.config.get("sheet_name", "")
            
            # Sayfa bul
            try: valid_sheet = wb.Sheets(target_sheet)
            except: valid_sheet = wb.ActiveSheet
                
            last_row = valid_sheet.Cells(valid_sheet.Rows.Count, 1).End(-4162).Row
            
            # Sadece gerekli sütunları oku (optimizasyon)
            # En sağdaki sütunu bul
            max_col_idx = 0
            for suffix, col_letter in self.dynamic_params:
                if col_letter:
                    col_idx = col2num(col_letter)
                    if col_idx > max_col_idx:
                        max_col_idx = col_idx
            
            # A sütunu (ID) + en sağdaki sütun + biraz buffer
            end_col = num2col(max(max_col_idx, 26))  # En az Z, gerekirse daha fazla
            
            # Excel verisini komple belleğe al (Hız için)
            raw_data = valid_sheet.Range(f"A2:{end_col}{last_row}").Value 
            # raw_data tuple of tuples döner. raw_data[satir_idx][sutun_idx]
            
            total_rows = len(raw_data) if raw_data else 0
            self.app.after(0, self.app.update_max_progress, total_rows)
            
            updates = 0
            errors = 0
            
            # --- PARAMETRE MAP --- (önceden hesapla)
            param_map = []
            for suffix, col_letter in self.dynamic_params:
                if not col_letter: continue
                col_idx = col2num(col_letter) - 1  # Python 0-based index
                param_map.append((suffix, col_idx))
            
            # CATIA bağlantısı (eğer gerekiyorsa)
            catia = None
            try:
                catia = win32com.client.GetActiveObject("CATIA.Application")
                # CATIA optimizasyonları
                if catia:
                    # Screen refresh'i kapat (batch update için)
                    # catia.DisplayFileAlerts = False
                    pass
            except:
                pass  # CATIA yoksa devam et
            
            # Batch update için değişiklikleri topla
            batch_size = 50  # Her 50 satırda bir UI güncelle
            last_update = 0
            
            # DÖNGÜ (optimize edilmiş)
            if raw_data:
                # Tuple yerine list'e çevir (daha hızlı erişim)
                data_list = list(raw_data) if isinstance(raw_data, tuple) else raw_data
                
                for i, row in enumerate(data_list):
                    if not self.running: break
                    
                    # ID Okuma (A Sütunu -> index 0)
                    try: 
                        id_val = row[0] if len(row) > 0 else None
                    except: 
                        continue
                    
                    if not id_val: continue
                    
                    # ID Formatlama (optimize edilmiş)
                    if isinstance(id_val, float) and id_val.is_integer():
                        id_str = str(int(id_val))
                    else:
                        id_str = str(id_val).strip()
                    
                    if not id_str: continue
                    
                    # Dinamik Parametreleri Güncelle
                    row_updates = 0
                    for suffix, col_idx in param_map:
                        if col_idx < len(row):
                            val = row[col_idx]
                            if val is not None and val != "":
                                # CATIA Parametre Adı: ID + Suffix
                                full_name = id_str + suffix
                                
                                # --- CATIA YAZMA KODU ---
                                try:
                                    # Değeri doğrula
                                    validated_value = validate_parameter_value(val, full_name)
                                    
                                    if catia:
                                        # CATIA'ya parametre yaz
                                        try:
                                            doc = catia.ActiveDocument
                                            part = doc.Part
                                            param = part.Parameters.Item(full_name)
                                            param.Value = validated_value
                                            # Parametreyi güncelle
                                            part.Update()
                                            APP_LOGGER.debug(f"{full_name} = {validated_value}")
                                        except Exception as catia_err:
                                            APP_LOGGER.error(f"CATIA parametresi yazılamadı ({full_name}): {catia_err}")
                                            raise
                                    
                                    updates += 1
                                    row_updates += 1
                                except ValueError as ve:
                                    errors += 1
                                    APP_LOGGER.warning(f"Doğrulama hatası - Satır {i+2}: {ve}")
                                    if errors <= 10:
                                        self.app.after(0, self.app.log, f"Satır {i+2}: {str(ve)}", "error")
                                except Exception as e:
                                    errors += 1
                                    APP_LOGGER.error(f"Beklenmeyen hata - Satır {i+2}: {e}\n{traceback.format_exc()}")
                                    if errors <= 10:
                                        self.app.after(0, self.app.log, f"Satır {i+2}: {full_name} = {str(e)}", "error")
                    
                    # Batch UI güncelleme (her N satırda bir veya son satır)
                    if (i + 1 - last_update >= batch_size) or (i + 1 == total_rows):
                        self.app.after(0, self.app.update_stats, i+1, updates, errors)
                        last_update = i + 1

            # Excel'i kapat
            wb.Close(False)
            excel.Quit()
            
            # Son güncelleme
            self.app.after(0, self.app.update_stats, total_rows, updates, errors)
            self.app.after(0, self.app.finish_process)

        except Exception as e:
            APP_LOGGER.critical(f"Kritik hata (real_process): {e}\n{traceback.format_exc()}")
            self.app.after(0, self.app.log, f"KRİTİK HATA: {e}", "error")
            
            # Cleanup - Excel'i güvenli şekilde kapat
            try:
                if 'excel' in locals() and excel:
                    APP_LOGGER.info("Excel kapatılıyor...")
                    if 'wb' in locals():
                        wb.Close(False)
                    excel.Quit()
                    APP_LOGGER.info("Excel kapatıldı")
            except Exception as cleanup_err:
                APP_LOGGER.error(f"Excel cleanup hatası: {cleanup_err}")
            
            # CATIA'yı serbest bırak
            try:
                if 'catia' in locals() and catia:
                    del catia
            except:
                pass
            
            # İşlemi bitir
            self.app.after(0, self.app.finish_process)

    def run_with_openpyxl(self):
        """openpyxl ile hızlı Excel okuma ve işleme"""
        try:
            APP_LOGGER.info(f"openpyxl ile dosya açılıyor: {self.excel_path}")
            sheet_name = self.config.get("sheet_name", None)
            
            # Excel'i oku
            data, sheets = read_excel_openpyxl(self.excel_path, sheet_name)
            
            total_rows = len(data)
            self.app.after(0, self.app.update_max_progress, total_rows)
            APP_LOGGER.info(f"Toplam {total_rows} satır okundu")
            
            updates = 0
            errors = 0
            
            # Parametre mapping'i hazırla
            param_map = []
            for suffix, col_letter in self.dynamic_params:
                if not col_letter:
                    continue
                col_idx = col2num(col_letter) - 1  # Python 0-based index
                param_map.append((suffix, col_idx))
            
            APP_LOGGER.info(f"Parametre mapping: {param_map}")
            
            # CATIA bağlantısı
            catia = None
            try:
                if WIN32COM_AVAILABLE:
                    import win32com.client
                    catia = win32com.client.GetActiveObject("CATIA.Application")
                    APP_LOGGER.info("CATIA bağlantısı başarılı")
                else:
                    APP_LOGGER.warning("win32com yok, CATIA'ya yazılamayacak (sadece simülasyon)")
            except Exception as catia_err:
                APP_LOGGER.warning(f"CATIA bağlanamadı: {catia_err}")
                self.app.after(0, self.app.log, "CATIA bağlanamadı - sadece simülasyon modu", "error")
            
            # Batch güncelleme
            batch_size = 50
            last_update = 0
            
            # Satırları işle
            for i, row in enumerate(data):
                if not self.running:
                    APP_LOGGER.info("İşlem kullanıcı tarafından durduruldu")
                    break
                
                # ID oku (ilk sütun)
                id_val = row[0] if len(row) > 0 else None
                if not id_val:
                    continue
                
                # ID formatla
                if isinstance(id_val, (int, float)):
                    id_str = str(int(id_val)) if float(id_val).is_integer() else str(id_val)
                else:
                    id_str = str(id_val).strip()
                
                if not id_str:
                    continue
                
                # Parametreleri güncelle
                row_updates = 0
                for suffix, col_idx in param_map:
                    if col_idx < len(row):
                        val = row[col_idx]
                        if val is not None and val != "":
                            full_name = id_str + suffix
                            
                            try:
                                # Değeri doğrula
                                validated_value = validate_parameter_value(val, full_name)
                                
                                if catia:
                                    # CATIA'ya yaz
                                    try:
                                        doc = catia.ActiveDocument
                                        part = doc.Part
                                        param = part.Parameters.Item(full_name)
                                        param.Value = validated_value
                                        part.Update()
                                        APP_LOGGER.debug(f"{full_name} = {validated_value}")
                                    except Exception as catia_err:
                                        APP_LOGGER.error(f"CATIA yazma hatası ({full_name}): {catia_err}")
                                        raise
                                
                                updates += 1
                                row_updates += 1
                                
                            except ValueError as ve:
                                errors += 1
                                APP_LOGGER.warning(f"Doğrulama hatası - Satır {i+2}: {ve}")
                                if errors <= 10:
                                    self.app.after(0, self.app.log, f"Satır {i+2}: {str(ve)}", "error")
                            except Exception as e:
                                errors += 1
                                APP_LOGGER.error(f"Hata - Satır {i+2}: {e}\n{traceback.format_exc()}")
                                if errors <= 10:
                                    self.app.after(0, self.app.log, f"Satır {i+2}: {full_name} = {str(e)}", "error")
                
                # Batch UI güncelleme
                if (i + 1 - last_update >= batch_size) or (i + 1 == total_rows):
                    self.app.after(0, self.app.update_stats, i+1, updates, errors)
                    last_update = i + 1
            
            # Son güncelleme
            self.app.after(0, self.app.update_stats, total_rows, updates, errors)
            self.app.after(0, self.app.finish_process)
            APP_LOGGER.info(f"İşlem tamamlandı - Başarılı: {updates}, Hata: {errors}")
            
        except Exception as e:
            APP_LOGGER.error(f"Kritik hata (openpyxl): {e}\n{traceback.format_exc()}")
            self.app.after(0, self.app.log, f"KRİTİK HATA: {e}", "error")
            self.app.after(0, self.app.finish_process)
    
    def stop(self):
        self.running = False
        APP_LOGGER.info("Durdurma talebi alındı")

# ==========================================
# EXCEL PREVIEW & ANALİZ
# ==========================================
class ExcelPreviewLoader(threading.Thread):
    def __init__(self, app, path):
        super().__init__()
        self.app = app
        self.path = path
        self.daemon = True

    def run(self):
        try:
            data_preview = []
            sheets = []
            
            if TEST_MODE:
                time.sleep(0.5)
                sheets = ["Visualization Data", "Sheet2"]
                # Fake Data: Header + 5 Rows
                data_preview = [
                    ["ID", "Th", "Height", "Mat", "X", "Y", "Z", "Q", "W", "E", "P1", "D1", "P2", "D2"],
                    ["101", "5.0", "20.0", "AL", "10", "20", "30", "-", "-", "-", "1.5", "0.5", "2.0", "0.8"],
                    ["102", "5.2", "21.0", "AL", "12", "22", "32", "-", "-", "-", "1.6", "0.6", "2.1", "0.9"],
                    ["103", "5.4", "22.0", "ST", "14", "24", "34", "-", "-", "-", "1.7", "0.7", "2.2", "1.0"],
                    ["104", "5.6", "23.0", "ST", "16", "26", "36", "-", "-", "-", "1.8", "0.8", "2.3", "1.1"],
                ]
            elif OPENPYXL_AVAILABLE:
                # openpyxl ile önizleme (daha hızlı)
                APP_LOGGER.info("openpyxl ile önizleme yükleniyor...")
                data_preview, sheets = read_excel_preview_openpyxl(self.path, max_rows=10)
            elif WIN32COM_AVAILABLE:
                # Fallback: win32com ile önizleme
                APP_LOGGER.info("win32com ile önizleme yükleniyor...")
                import win32com.client
                import pythoncom
                pythoncom.CoInitialize()
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                excel.ScreenUpdating = False
                excel.DisplayAlerts = False
                excel.EnableEvents = False
                excel.Calculation = -4135  # xlCalculationManual
                
                wb = excel.Workbooks.Open(self.path, ReadOnly=True)
                
                # Sayfaları al
                sheets = [s.Name for s in wb.Sheets]
                
                # İlk sayfadan önizleme al (İlk 10 satır, ilk 15 sütun)
                ws = wb.Sheets(1)
                vals = ws.Range("A1:O10").Value
                # Tuple to List (batch conversion)
                if vals:
                    if isinstance(vals, tuple):
                        data_preview = [list(row) if row else [] for row in vals]
                    else:
                        data_preview = [list(row) if row else [] for row in [vals]]
                
                wb.Close(False)
                excel.Quit()
            else:
                raise ImportError("Excel okuma için openpyxl veya pywin32 gerekli")

            self.app.after(0, self.app.update_ui_with_excel_data, sheets, data_preview)

        except Exception as e:
            APP_LOGGER.error(f"Önizleme hatası: {e}\n{traceback.format_exc()}")
            self.app.after(0, self.app.log, f"Önizleme hatası: {e}", "error")
            
            # Hata detayını göster
            error_msg = f"Excel önizlemesi yüklenemedi:\n\n{str(e)}\n\n"
            if not OPENPYXL_AVAILABLE and not WIN32COM_AVAILABLE:
                error_msg += "Çözüm: 'pip install openpyxl' veya 'pip install pywin32' ile gerekli kütüphaneleri yükleyin."
            
            self.app.after(0, messagebox.showerror, "Önizleme Hatası", error_msg)

# ==========================================
# ÖZEL BUTON WIDGET'LARI
# ==========================================
class ModernButton(ctk.CTkFrame):
    """Modern hover efektli buton"""
    def __init__(self, parent, text="", command=None, icon="→", fg_color="#3B8ED0", 
                 hover_color="#2E7BD6", text_color="white", width=120, height=40):
        super().__init__(parent, fg_color="transparent", width=width, height=height)
        self.command = command
        self.fg_color = fg_color
        self.hover_color = hover_color
        self.text_color = text_color
        self.icon = icon
        self.text = text
        
        self.button_canvas = ctk.CTkCanvas(self, width=width, height=height, 
                                           highlightthickness=0, bg="#1a1a1a")
        self.button_canvas.pack(fill="both", expand=True)
        
        # Canvas boyutunu zorla ayarla ve çiz
        self.button_canvas.update_idletasks()
        self.after(10, self.draw_button)
        self.button_canvas.configure(cursor="hand2")
        self.button_canvas.bind("<Enter>", self.on_enter)
        self.button_canvas.bind("<Leave>", self.on_leave)
        self.button_canvas.bind("<Button-1>", self.on_click)
        self.button_canvas.bind("<ButtonRelease-1>", self.on_release)
        self.button_canvas.bind("<Configure>", lambda e: self.draw_button())
        
    def draw_button(self, hover=False):
        self.button_canvas.delete("all")
        w = self.button_canvas.winfo_width()
        h = self.button_canvas.winfo_height()
        if w < 10 or h < 10:
            return
            
        color = self.hover_color if hover else self.fg_color
        
        # Buton arka planı
        self.button_canvas.create_rectangle(2, 2, w-2, h-2, fill=color, 
                                           outline=color, width=0, tags="bg")
        
        # Metin
        self.button_canvas.create_text(w//2 - 8, h//2, text=self.text, 
                                      fill=self.text_color, font=("Roboto", 12, "bold"),
                                      anchor="e", tags="text")
        
        # Ok ikonu
        self.button_canvas.create_text(w//2 + 8, h//2, text=self.icon, 
                                      fill=self.text_color, font=("Arial", 14),
                                      anchor="w", tags="icon")
        
    def on_enter(self, e):
        self.draw_button(hover=True)
        
    def on_leave(self, e):
        self.draw_button(hover=False)
        
    def on_click(self, e):
        if self.command:
            self.command()
            
    def on_release(self, e):
        pass
        
    def set_state(self, state):
        """Buton state'ini ayarlar"""
        if state == "disabled":
            self.button_canvas.configure(cursor="")
            self.button_canvas.unbind("<Enter>")
            self.button_canvas.unbind("<Leave>")
            self.button_canvas.unbind("<Button-1>")
            self.draw_button_disabled()
        else:
            self.button_canvas.configure(cursor="hand2")
            self.button_canvas.bind("<Enter>", self.on_enter)
            self.button_canvas.bind("<Leave>", self.on_leave)
            self.button_canvas.bind("<Button-1>", self.on_click)
            self.draw_button()
            
    def draw_button_disabled(self):
        """Disabled durumda çiz"""
        self.button_canvas.delete("all")
        w = self.button_canvas.winfo_width()
        h = self.button_canvas.winfo_height()
        if w < 10 or h < 10:
            return
            
        # Gri buton
        self.button_canvas.create_rectangle(2, 2, w-2, h-2, fill="#555", 
                                           outline="#555", width=0, tags="bg")
        
        # Metin
        self.button_canvas.create_text(w//2 - 8, h//2, text=self.text, 
                                      fill="#999", font=("Roboto", 12, "bold"),
                                      anchor="e", tags="text")
        
        # Ok ikonu
        self.button_canvas.create_text(w//2 + 8, h//2, text=self.icon, 
                                      fill="#999", font=("Arial", 14),
                                      anchor="w", tags="icon")
                                      
    def update_text(self, text):
        """Buton metnini güncelle"""
        self.text = text
        self.draw_button()
        
    def update_color(self, fg_color, hover_color=None):
        """Buton rengini güncelle"""
        self.fg_color = fg_color
        if hover_color:
            self.hover_color = hover_color
        self.draw_button()

class IconButton(ctk.CTkFrame):
    """İkonlu buton (Download, Delete vb.)"""
    def __init__(self, parent, icon="⬇", command=None, tooltip="", 
                 fg_color="#2b2b2b", hover_color="#3b3b3b", size=40):
        super().__init__(parent, fg_color="transparent", width=size, height=size)
        self.icon = icon
        self.command = command
        self.tooltip = tooltip
        self.fg_color = fg_color
        self.hover_color = hover_color
        self.size = size
        
        self.button_canvas = ctk.CTkCanvas(self, width=size, height=size,
                                         highlightthickness=0, bg="#1a1a1a")
        self.button_canvas.pack(fill="both", expand=True)
        
        # Canvas boyutunu zorla ayarla ve çiz
        self.button_canvas.update_idletasks()
        self.after(10, self.draw_button)
        self.button_canvas.configure(cursor="hand2")
        self.button_canvas.bind("<Enter>", self.on_enter)
        self.button_canvas.bind("<Leave>", self.on_leave)
        self.button_canvas.bind("<Button-1>", self.on_click)
        self.button_canvas.bind("<Configure>", lambda e: self.draw_button())
        
        # Tooltip
        self.tooltip_label = None
        
    def draw_button(self, hover=False):
        self.button_canvas.delete("all")
        w = self.size
        h = self.size
        color = self.hover_color if hover else self.fg_color
        
        # Yuvarlak buton
        self.button_canvas.create_oval(4, 4, w-4, h-4, fill=color, 
                                      outline="#444", width=1, tags="bg")
        
        # İkon (Unicode karakterler)
        icon_map = {
            "⬇": "⬇", "🗑": "🗑", "↑": "↑", "→": "→",
            "download": "⬇", "delete": "🗑", "up": "↑"
        }
        icon_char = icon_map.get(self.icon, self.icon)
        self.button_canvas.create_text(w//2, h//2, text=icon_char,
                                      fill="white", font=("Arial", 16),
                                      tags="icon")
        
    def on_enter(self, e):
        self.draw_button(hover=True)
        if self.tooltip:
            self.show_tooltip(e)
            
    def on_leave(self, e):
        self.draw_button(hover=False)
        self.hide_tooltip()
        
    def on_click(self, e):
        if self.command:
            self.command()
            
    def show_tooltip(self, e):
        if self.tooltip_label:
            self.tooltip_label.destroy()
        self.tooltip_label = ctk.CTkToplevel(self)
        self.tooltip_label.overrideredirect(True)
        self.tooltip_label.attributes("-topmost", True)
        label = ctk.CTkLabel(self.tooltip_label, text=self.tooltip, 
                            fg_color="#333", text_color="white",
                            corner_radius=5, padx=8, pady=4)
        label.pack()
        x = e.x_root + 10
        y = e.y_root - 30
        self.tooltip_label.geometry(f"+{x}+{y}")
        
    def hide_tooltip(self):
        if self.tooltip_label:
            self.tooltip_label.destroy()
            self.tooltip_label = None
            
    def set_state(self, state):
        """Buton state'ini ayarlar"""
        if state == "disabled":
            self.button_canvas.configure(cursor="")
            self.button_canvas.unbind("<Enter>")
            self.button_canvas.unbind("<Leave>")
            self.button_canvas.unbind("<Button-1>")
            self.draw_button_disabled()
        else:
            self.button_canvas.configure(cursor="hand2")
            self.button_canvas.bind("<Enter>", self.on_enter)
            self.button_canvas.bind("<Leave>", self.on_leave)
            self.button_canvas.bind("<Button-1>", self.on_click)
            self.draw_button()
            
    def draw_button_disabled(self):
        """Disabled durumda çiz"""
        self.button_canvas.delete("all")
        w = self.size
        h = self.size
        # Gri buton
        self.button_canvas.create_oval(4, 4, w-4, h-4, fill="#555", 
                                      outline="#444", width=1, tags="bg")
        # Gri ikon
        icon_map = {"⬇": "⬇", "🗑": "🗑", "↑": "↑", "→": "→",
                   "download": "⬇", "delete": "🗑", "up": "↑"}
        icon_char = icon_map.get(self.icon, self.icon)
        self.button_canvas.create_text(w//2, h//2, text=icon_char,
                                      fill="#999", font=("Arial", 16),
                                      tags="icon")

class LearnMoreButton(ctk.CTkFrame):
    """Circle animasyonlu Learn More butonu"""
    def __init__(self, parent, text="Learn More", command=None, 
                 fg_color="#3B8ED0", hover_color="#2E7BD6"):
        super().__init__(parent, fg_color="transparent")
        self.command = command
        self.fg_color = fg_color
        self.hover_color = hover_color
        self.text = text
        self.circle_radius = 0
        
        self.button_canvas = ctk.CTkCanvas(self, width=140, height=40,
                                         highlightthickness=0, bg="#1a1a1a")
        self.button_canvas.pack(fill="both", expand=True)
        
        # Canvas boyutunu zorla ayarla ve çiz
        self.button_canvas.update_idletasks()
        self.after(10, self.draw_button)
        self.button_canvas.configure(cursor="hand2")
        self.button_canvas.bind("<Enter>", self.on_enter)
        self.button_canvas.bind("<Leave>", self.on_leave)
        self.button_canvas.bind("<Button-1>", self.on_click)
        self.button_canvas.bind("<Configure>", lambda e: self.draw_button())
        
    def draw_button(self, hover=False):
        self.button_canvas.delete("all")
        w = self.button_canvas.winfo_width()
        h = self.button_canvas.winfo_height()
        if w < 10 or h < 10:
            return
            
        color = self.hover_color if hover else self.fg_color
        
        # Circle animasyonu
        if hover and self.circle_radius < 50:
            self.circle_radius += 2
            self.button_canvas.create_oval(w//2 - self.circle_radius, h//2 - self.circle_radius,
                                          w//2 + self.circle_radius, h//2 + self.circle_radius,
                                          fill=color, outline="", tags="circle")
            self.after(20, lambda: self.draw_button(hover=True) if hover else None)
        elif not hover:
            self.circle_radius = 0
        
        # Buton arka planı
        self.button_canvas.create_rectangle(2, 2, w-2, h-2, fill=color, 
                                           outline=color, width=0, tags="bg")
        
        # Metin
        self.button_canvas.create_text(w//2, h//2, text=self.text,
                                      fill="white", font=("Roboto", 11, "bold"),
                                      tags="text")
        
        # Ok ikonu
        self.button_canvas.create_text(w//2 + 50, h//2, text="→",
                                      fill="white", font=("Arial", 12),
                                      tags="icon")
        
    def on_enter(self, e):
        self.draw_button(hover=True)
        
    def on_leave(self, e):
        self.circle_radius = 0
        self.draw_button(hover=False)
        
    def on_click(self, e):
        if self.command:
            self.command()
            
    def update_text(self, text):
        """Buton metnini güncelle"""
        self.text = text
        self.draw_button()

# ==========================================
# ARAYÜZ (GUI)
# ==========================================
class AutomationSuite(ctk.CTk):
    def __init__(self):
        super().__init__()
        
        APP_LOGGER.info("Uygulama başlatılıyor...")
        
        # Config
        self.config = {"sheet_name": "", "always_on_top": False}
        self.param_rows = [] 

        # Tema ve Pencere
        ctk.set_appearance_mode("Dark")
        ctk.set_default_color_theme("dark-blue") # Mavi vurgular
        
        # Renk Paleti (Custom Colors)
        self.colors = {
            "bg": "#1a1a1a",            # Ana arka plan (Daha koyu)
            "panel": "#2b2b2b",         # Paneller
            "card": "#333333",          # Parametre kartları
            "accent": "#3B8ED0",        # Mavi vurgu
            "danger": "#cf6679",        # Kırmızı (Silme)
            "text_gray": "#a1a1aa",     # Gri yazı
            "grid_header": "#404040",   # Tablo başlığı
            "grid_row_even": "#262626", # Tablo satır 1
            "grid_row_odd": "#1f1f1f"   # Tablo satır 2
        }
        self.title("CATIA Automation Suite v4.5 Pro")
        self.geometry("1100x750")
        
        # Değişkenler (setup_monitor'dan önce tanımlanmalı)
        self.log_entries = []
        self.show_errors_only = False
        self.preview_table = None
        self.tab_titles = {
            "monitor": "  Monitör  ",
            "settings": "  Ayarlar & Önizleme  "
        }
        self.current_run_has_error = False
        self.selected_file = None
        self.worker = None
        self.total_work = 1
        self.start_time = 0
        
        # --- TAB YAPISI ---
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)
        
        self.tab_view = ctk.CTkTabview(self, corner_radius=15, fg_color=self.colors["bg"])
        self.tab_view.grid(row=0, column=0, padx=20, pady=20, sticky="nsew")
        
        self.tab_monitor = self.tab_view.add("  Monitör  ")
        self.tab_settings = self.tab_view.add("  Ayarlar & Önizleme  ")
        self.setup_monitor()
        self.setup_settings()
        self.update_tab_headers(False)
        
        # Klavye kısayolları
        self.bind("<Control-s>", lambda e: self.save_profile())
        self.bind("<Control-o>", lambda e: self.load_profile())
        self.bind("<Control-r>", lambda e: self.start_process() if self.btn_run.cget("state") == "normal" else None)
        self.bind("<Control-w>", lambda e: self.stop_process() if self.btn_stop.cget("state") == "normal" else None)
        self.bind("<Control-e>", lambda e: self.export_html_report())
        self.bind("<Control-t>", lambda e: self.create_excel_template())
        self.bind("<Control-f>", lambda e: self.select_file())
        
        # Uygulama kapatılırken logger'ı kapat
        self.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        # Varsayılan profili yükle
        self.load_default_profile()

    # ------------------------------
    # MONİTÖR TABI (Aynı kalabilir veya ufak makyaj)
    # ------------------------------
    def setup_monitor(self):
        self.tab_monitor.grid_columnconfigure(0, weight=1)
        self.tab_monitor.grid_rowconfigure(3, weight=1)
        # Üst Panel
        frame_top = ctk.CTkFrame(self.tab_monitor, fg_color="transparent")
        frame_top.grid(row=0, column=0, padx=10, pady=10, sticky="ew")
        
        # Modern butonlar - CTkButton ile hover efektleri
        self.btn_file = ctk.CTkButton(frame_top, text="📂 Excel Dosyası Seç →", 
                                     command=self.select_file,
                                     fg_color="#3B8ED0", hover_color="#2E7BD6",
                                     width=180, height=40, font=("Roboto", 13, "bold"),
                                     corner_radius=8)
        self.btn_file.pack(side="left", padx=5)
        
        self.lbl_file = ctk.CTkLabel(frame_top, text="Henüz dosya seçilmedi", text_color=self.colors["text_gray"], font=("Roboto", 12))
        self.lbl_file.pack(side="left", padx=15)
        
        self.status_badge = ctk.CTkLabel(frame_top, text="Durum: Beklemede", text_color="#F1C40F", font=("Roboto", 12, "bold"))
        self.status_badge.pack(side="right", padx=10)
        
        self.btn_run = ctk.CTkButton(frame_top, text="İŞLEMİ BAŞLAT ▶", 
                                    command=self.start_process,
                                    fg_color="#2ECC71", hover_color="#25a25a",
                                    width=180, height=40, font=("Roboto", 13, "bold"),
                                    corner_radius=8, state="disabled")
        self.btn_run.pack(side="right", padx=5)

        summary_bar = ctk.CTkFrame(self.tab_monitor, fg_color="transparent")
        summary_bar.grid(row=1, column=0, padx=10, pady=(0,10), sticky="ew")
        summary_bar.grid_columnconfigure((0,1), weight=1)
        self.lbl_summary = ctk.CTkLabel(summary_bar, text="Dosya: - | Sayfa: -", text_color=self.colors["text_gray"])
        self.lbl_summary.grid(row=0, column=0, sticky="w")
        self.lbl_sheet_hint = ctk.CTkLabel(summary_bar, text="Parametre eşleşmelerini Ayarlar sekmesinden kontrol edin.", text_color="#888")
        self.lbl_sheet_hint.grid(row=0, column=1, sticky="e")
        # İstatistik Kartları
        frame_stats = ctk.CTkFrame(self.tab_monitor, fg_color=self.colors["panel"], corner_radius=10)
        frame_stats.grid(row=2, column=0, padx=10, pady=10, sticky="ew")
        frame_stats.columnconfigure((0,1,2,3,4,5), weight=1)
        
        self.lbl_time = self._card(frame_stats, 0, "Geçen Süre")
        self.lbl_eta = self._card(frame_stats, 1, "Tahmini Bitiş")
        self.lbl_upd = self._card(frame_stats, 2, "Başarılı", "#2ECC71")
        self.lbl_err = self._card(frame_stats, 3, "Hatalar", "#cf6679")
        self.lbl_success_rate = self._card(frame_stats, 4, "Başarı Oranı", "#3B8ED0")
        self.lbl_avg_time = self._card(frame_stats, 5, "Ort. Süre/İtem", "#F39C12")
        # Log
        self.log_box = ctk.CTkTextbox(self.tab_monitor, font=("Consolas", 12), state="disabled", 
                                      fg_color="#111", text_color="#eee", corner_radius=10)
        self.log_box.grid(row=3, column=0, padx=10, pady=10, sticky="nsew")
        # Progress - Hamster Wheel Animation
        progress_frame = ctk.CTkFrame(self.tab_monitor, fg_color="transparent")
        progress_frame.grid(row=4, column=0, padx=10, pady=(5,10), sticky="ew")
        progress_frame.grid_columnconfigure(0, weight=1)
        
        # Hamster wheel canvas
        self.hamster_canvas = ctk.CTkCanvas(progress_frame, height=80, bg="#1a1a1a", highlightthickness=0)
        self.hamster_canvas.grid(row=0, column=0, sticky="ew", padx=10)
        
        def on_canvas_configure(e):
            self.draw_hamster_wheel()
        self.hamster_canvas.bind("<Configure>", on_canvas_configure)
        
        # İlk çizimi başlat (canvas boyutlandıktan sonra)
        self.after(200, self.initial_draw_hamster)
        
        # Progress bar (gizli, sadece değer tutmak için)
        self.progress = ctk.CTkProgressBar(progress_frame, height=15, corner_radius=8)
        self.progress.grid(row=1, column=0, sticky="ew", pady=(5,0))
        self.progress.set(0)
        
        # Animasyon değişkenleri
        self.hamster_angle = 0
        self.hamster_animation_id = None

        log_actions = ctk.CTkFrame(self.tab_monitor, fg_color="transparent")
        log_actions.grid(row=5, column=0, padx=10, pady=(0,10), sticky="ew")
        log_actions.grid_columnconfigure((0,1,2), weight=1)

        # Modern butonlar - CTkButton ile
        self.btn_toggle_errors = ctk.CTkButton(log_actions, text="Sadece Hatalar",
                                               command=self.toggle_error_filter,
                                               fg_color=self.colors["panel"],
                                               hover_color=self.colors["card"],
                                               text_color=self.colors["accent"],
                                               height=34, corner_radius=8)
        self.btn_toggle_errors.grid(row=0, column=0, padx=5, sticky="ew")

        self.btn_save_log = ctk.CTkButton(log_actions, text="⬇ Logu Kaydet",
                                         command=self.save_log,
                                         fg_color="#3B8ED0", hover_color="#2E7BD6",
                                         height=34, corner_radius=8)
        self.btn_save_log.grid(row=0, column=1, padx=5, sticky="ew")

        self.btn_stop = ctk.CTkButton(log_actions, text="🗑 DURDUR",
                                      command=self.stop_process,
                                      fg_color=self.colors["danger"],
                                      hover_color="#E74C3C",
                                      height=34, corner_radius=8, state="disabled")
        self.btn_stop.grid(row=0, column=2, padx=5, sticky="ew")
        
        # Profil ve Araçlar
        tools_actions = ctk.CTkFrame(self.tab_monitor, fg_color="transparent")
        tools_actions.grid(row=6, column=0, padx=10, pady=(5,10), sticky="ew")
        tools_actions.grid_columnconfigure((0,1,2,3), weight=1)
        
        self.btn_save_profile = ctk.CTkButton(tools_actions, text="💾 Kaydet (Ctrl+S)",
                                             command=self.save_profile,
                                             fg_color=self.colors["panel"],
                                             hover_color=self.colors["card"],
                                             height=32, corner_radius=8)
        self.btn_save_profile.grid(row=0, column=0, padx=3, sticky="ew")
        
        self.btn_load_profile = ctk.CTkButton(tools_actions, text="📂 Yükle (Ctrl+O)",
                                             command=self.load_profile,
                                             fg_color=self.colors["accent"],
                                             hover_color="#2E7BD6",
                                             height=32, corner_radius=8)
        self.btn_load_profile.grid(row=0, column=1, padx=3, sticky="ew")
        
        self.btn_export_report = ctk.CTkButton(tools_actions, text="📄 Rapor (Ctrl+E)",
                                              command=self.export_html_report,
                                              fg_color="#9B59B6",
                                              hover_color="#8E44AD",
                                              height=32, corner_radius=8)
        self.btn_export_report.grid(row=0, column=2, padx=3, sticky="ew")
        
        self.btn_create_template = ctk.CTkButton(tools_actions, text="📋 Şablon (Ctrl+T)",
                                                command=self.create_excel_template,
                                                fg_color="#16A085",
                                                hover_color="#138D75",
                                                height=32, corner_radius=8)
        self.btn_create_template.grid(row=0, column=3, padx=3, sticky="ew")

        self.set_running_state(False)
        self.update_summary_label()
        self.update_error_button_text()

    def _card(self, parent, col, title, color="white"):
        f = ctk.CTkFrame(parent, fg_color="transparent")
        f.grid(row=0, column=col, padx=10, pady=10)
        ctk.CTkLabel(f, text=title, font=("Roboto", 11), text_color=self.colors["text_gray"]).pack()
        l = ctk.CTkLabel(f, text="--", font=("Roboto", 20, "bold"), text_color=color)
        l.pack()
        return l

    # ------------------------------
    # AYARLAR & ÖNİZLEME TABI (YENİ TASARIM)
    # ------------------------------
    def setup_settings(self):
        # Grid: Sol (4 birim), Sağ (6 birim)
        self.tab_settings.grid_columnconfigure(0, weight=2) # Sol Panel
        self.tab_settings.grid_columnconfigure(1, weight=3) # Sağ Panel
        self.tab_settings.grid_rowconfigure(0, weight=1)
        # --- SOL PANEL: PARAMETRELER ---
        left_container = ctk.CTkFrame(self.tab_settings, fg_color="transparent")
        left_container.grid(row=0, column=0, padx=(0, 10), pady=10, sticky="nsew")
        
        # Başlık Alanı
        header_lbl = ctk.CTkLabel(left_container, text="Parametre Eşleştirme", 
                                 font=("Roboto", 18, "bold"), text_color="white", anchor="w")
        header_lbl.pack(fill="x", pady=(0, 15))
        # Scrollable Parametre Listesi
        self.scroll_params = ctk.CTkScrollableFrame(left_container, label_text="", fg_color="transparent")
        self.scroll_params.pack(fill="both", expand=True, pady=(0, 10))
        # Varsayılan Parametreler
        self.default_params = [
            ("Thickness", "B"), ("H", "C"),
            ("P1", "K"), ("D1", "L"),
            ("P2", "M"), ("D2", "N")
        ]
        for name, col in self.default_params:
            self.add_param_row(name, col)
        # Ekleme Butonu
        self.btn_add_param = ctk.CTkButton(left_container, text="+ Parametre Ekle", 
                               command=lambda: self.add_param_row("", ""), 
                               fg_color=self.colors["panel"], hover_color=self.colors["card"],
                               text_color=self.colors["accent"], font=("Roboto", 12, "bold"),
                               height=40, border_width=1, border_color=self.colors["card"])
        self.btn_add_param.pack(fill="x", pady=(0, 10))

        param_tools = ctk.CTkFrame(left_container, fg_color="transparent")
        param_tools.pack(fill="x", pady=(0, 15))
        param_tools.grid_columnconfigure((0,1,2), weight=1)

        self.btn_sort_params = ctk.CTkButton(
            param_tools,
            text="A→Z Sırala",
            command=self.sort_param_rows,
            fg_color=self.colors["panel"],
            hover_color=self.colors["card"],
            height=34
        )
        self.btn_sort_params.grid(row=0, column=0, padx=5, sticky="ew")

        self.btn_clear_params = ctk.CTkButton(
            param_tools,
            text="Sıfırla",
            command=self.clear_param_rows,
            fg_color=self.colors["danger"],
            height=34
        )
        self.btn_clear_params.grid(row=0, column=1, padx=5, sticky="ew")
        
        self.btn_templates = ctk.CTkButton(
            param_tools,
            text="📋 Şablon",
            command=self.show_template_menu,
            fg_color=self.colors["accent"],
            hover_color="#2E7BD6",
            height=34
        )
        self.btn_templates.grid(row=0, column=2, padx=5, sticky="ew")
        
        # Alt Ayarlar (Sayfa Seçimi)
        settings_card = ctk.CTkFrame(left_container, fg_color=self.colors["panel"], corner_radius=10)
        settings_card.pack(fill="x")
        
        ctk.CTkLabel(settings_card, text="Hedef Excel Sayfası", font=("Roboto", 12, "bold"), text_color=self.colors["text_gray"]).pack(anchor="w", padx=15, pady=(10,5))
        self.combo_sheet = ctk.CTkComboBox(settings_card, values=["Dosya Bekleniyor..."], height=35, font=("Roboto", 13), command=self.on_sheet_change)
        self.combo_sheet.pack(fill="x", padx=15, pady=(0, 15))
        # --- SAĞ PANEL: CANLI ÖNİZLEME ---
        right_container = ctk.CTkFrame(self.tab_settings, fg_color=self.colors["panel"], corner_radius=15)
        right_container.grid(row=0, column=1, padx=(10, 0), pady=10, sticky="nsew")
        
        # Başlık
        ctk.CTkLabel(right_container, text="Canlı Veri Önizleme", font=("Roboto", 16, "bold")).pack(pady=15)
        
        # Tablo Alanı (Scrollable)
        self.preview_box = ctk.CTkScrollableFrame(right_container, fg_color="#1e1e1e", corner_radius=10)
        self.preview_box.pack(fill="both", expand=True, padx=15, pady=(0, 15))
        
        # Empty State
        self.empty_state_frame = ctk.CTkFrame(self.preview_box, fg_color="transparent")
        self.empty_state_frame.pack(expand=True, fill="both", pady=100)
        ctk.CTkLabel(self.empty_state_frame, text="📊", font=("Arial", 40)).pack()
        ctk.CTkLabel(self.empty_state_frame, text="Önizleme için bir Excel dosyası seçin.", 
                     font=("Roboto", 14), text_color="gray").pack(pady=10)

    def add_param_row(self, default_name, default_col):
        """
        Daha şık, 'Kart' görünümlü parametre satırı.
        """
        # Kart Konteyner
        card = ctk.CTkFrame(self.scroll_params, fg_color=self.colors["card"], corner_radius=8)
        card.pack(fill="x", pady=4, padx=2)
        
        # İçerik için bir frame (pack kullanarak)
        content_frame = ctk.CTkFrame(card, fg_color="transparent")
        content_frame.pack(fill="x", padx=10, pady=10)
        
        # 1. Sütun Harfi (Badge Style)
        col_container = ctk.CTkFrame(content_frame, fg_color="transparent")
        col_container.pack(side="left", padx=(0, 8))
        
        ctk.CTkLabel(col_container, text="Sütun", font=("Arial", 9), text_color="gray").pack(anchor="w", pady=(0,2))
        col_entry = ctk.CTkEntry(col_container, width=45, height=32, 
                                 font=("Roboto", 13, "bold"), justify="center",
                                 fg_color=self.colors["panel"], 
                                 border_width=1, border_color=self.colors["card"],
                                 corner_radius=6)
        col_entry.insert(0, default_col)
        col_entry.pack()
        
        # Ok İkonu
        arrow_lbl = ctk.CTkLabel(content_frame, text="➜", text_color="#888", font=("Arial", 16))
        arrow_lbl.pack(side="left", padx=8)
        
        # 2. Parametre Adı (Suffix)
        name_container = ctk.CTkFrame(content_frame, fg_color="transparent")
        name_container.pack(side="left", fill="x", expand=True, padx=(0, 8))
        
        ctk.CTkLabel(name_container, text="CATIA Parametre Adı (Ek)", font=("Arial", 9), text_color="gray").pack(anchor="w", pady=(0,2))
        name_entry = ctk.CTkEntry(name_container, height=32, font=("Roboto", 12),
                                  placeholder_text="Örn: Thickness", 
                                  fg_color=self.colors["panel"],
                                  border_width=1, border_color=self.colors["card"],
                                  corner_radius=6)
        name_entry.insert(0, default_name)
        name_entry.pack(fill="x")
        
        # 3. Canlı ID Önizlemesi (Küçük bilgi)
        info_frame = ctk.CTkFrame(content_frame, fg_color="transparent")
        info_frame.pack(side="left", padx=(0, 5))
        ctk.CTkLabel(info_frame, text="→", font=("Arial", 12), text_color="#666").pack()
        info_lbl = ctk.CTkLabel(info_frame, text=f"ID{default_name}", 
                               font=("Consolas", 10, "bold"), 
                               text_color=self.colors["accent"])
        info_lbl.pack()
        
        # 4. Silme Butonu (Ghost Style)
        del_btn = ctk.CTkButton(content_frame, text="✕", width=32, height=32,
                                fg_color="transparent", text_color="#999",
                                hover_color=self.colors["danger"],
                                font=("Arial", 16, "bold"),
                                corner_radius=6,
                                command=lambda: self.delete_param_row(card))
        del_btn.pack(side="right", padx=(5, 0))
        
        # Referansları Sakla
        def normalize_col(event=None):
            raw = col_entry.get().upper()
            normalized = "".join([ch for ch in raw if ch.isalpha()])
            if raw != normalized:
                col_entry.delete(0, "end")
                col_entry.insert(0, normalized)
            if normalized:
                col_entry.configure(text_color="white")
            else:
                col_entry.configure(text_color=self.colors["danger"])

        def update_info(event=None):
            suf = name_entry.get().strip()
            info_lbl.configure(text=f"ID{suf}" if suf else "ID...")

        col_entry.bind("<KeyRelease>", normalize_col)
        name_entry.bind("<KeyRelease>", update_info)
        normalize_col()
        update_info()

        self.param_rows.append({
            "frame": card,
            "col": col_entry,
            "name": name_entry,
            "info": info_lbl,
            "delete_btn": del_btn
        })

    def delete_param_row(self, frame):
        frame.destroy()
        self.param_rows = [r for r in self.param_rows if r["frame"] != frame]

    def sort_param_rows(self):
        ordered = sorted(self.param_rows, key=lambda r: r["name"].get().strip().upper() or "~")
        for row in ordered:
            row["frame"].pack_forget()
            row["frame"].pack(fill="x", pady=4, padx=2)
        self.param_rows = ordered

    def clear_param_rows(self):
        for row in self.param_rows:
            row["frame"].destroy()
        self.param_rows = []
        for name, col in self.default_params:
            self.add_param_row(name, col)

    def set_controls_state(self, enabled: bool):
        state = "normal" if enabled else "disabled"
        combo_state = "normal" if enabled else "disabled"
        if hasattr(self, "btn_file"):
            self.btn_file.configure(state=state)
        if hasattr(self, "btn_add_param"):
            self.btn_add_param.configure(state=state)
        if hasattr(self, "btn_sort_params"):
            self.btn_sort_params.configure(state=state)
        if hasattr(self, "btn_clear_params"):
            self.btn_clear_params.configure(state=state)
        if hasattr(self, "combo_sheet"):
            self.combo_sheet.configure(state=combo_state)
        for row in self.param_rows:
            row["col"].configure(state=state)
            row["name"].configure(state=state)
            row["delete_btn"].configure(state=state)

    def update_tab_headers(self, running: bool):
        segmented = self.tab_view._segmented_button
        monitor_key = self.tab_titles["monitor"]
        settings_key = self.tab_titles["settings"]
        monitor_text = "  Monitör (Çalışıyor)  " if running else monitor_key
        settings_text = "  Ayarlar & Önizleme (Kilitli)  " if running else settings_key
        segmented._buttons_dict[monitor_key].configure(text=monitor_text)
        segmented._buttons_dict[settings_key].configure(text=settings_text)

    def set_running_state(self, running: bool):
        self.update_tab_headers(running)
        if running:
            self.status_badge.configure(text="Durum: Çalışıyor", text_color="#2ECC71")
        elif self.current_run_has_error:
            self.status_badge.configure(text="Durum: Hatalar Var", text_color=self.colors["danger"])
        else:
            self.status_badge.configure(text="Durum: Beklemede", text_color="#F1C40F")

    def update_summary_label(self):
        file_txt = os.path.basename(self.selected_file) if self.selected_file else "-"
        sheet_name = ""
        if hasattr(self, "combo_sheet"):
            sheet_name = self.combo_sheet.get() or "-"
        self.lbl_summary.configure(text=f"Dosya: {file_txt} | Sayfa: {sheet_name or '-'}")

    def on_sheet_change(self, value: str):
        self.config["sheet_name"] = value
        self.update_summary_label()

    def update_error_button_text(self):
        if not hasattr(self, "btn_toggle_errors"):
            return
        error_count = sum(1 for entry in self.log_entries if entry["type"] == "error")
        if self.show_errors_only:
            text = f"Tüm Kayıtları Göster ({error_count})"
        else:
            text = f"Sadece Hatalar ({error_count})"
        if hasattr(self, "btn_toggle_errors"):
            self.btn_toggle_errors.configure(text=text)

    # ------------------------------
    # İŞLEVLER
    # ------------------------------
    def select_file(self):
        was_topmost = bool(self.attributes("-topmost"))
        self.attributes("-topmost", False)
        path = ctk.filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx *.xlsm")])
        
        if path:
            # Excel validasyonu
            if not self.validate_excel_file(path):
                self.log("Geçersiz Excel dosyası! Lütfen .xlsx veya .xlsm formatında bir dosya seçin.", "error")
                self.attributes("-topmost", was_topmost)
                return
            
            self.selected_file = path
            self.lbl_file.configure(text=os.path.basename(path), text_color="white")
            self.btn_run.configure(state="normal")
            self.log(f"Dosya seçildi: {os.path.basename(path)}", "info")
            
            # Tabı Ayarlar'a geçir
            self.tab_view.set("  Ayarlar & Önizleme  ")
            self.log("Önizleme oluşturuluyor...", "info")
            ExcelPreviewLoader(self, path).start()
            self.update_summary_label()
        
        self.attributes("-topmost", was_topmost)
    
    def validate_excel_file(self, path):
        """Excel dosyasını doğrular"""
        try:
            # Dosya varlık kontrolü
            if not os.path.exists(path):
                self.log("Dosya bulunamadı!", "error")
                APP_LOGGER.error(f"Dosya bulunamadı: {path}")
                return False
            
            # Format kontrolü
            if not path.lower().endswith(('.xlsx', '.xlsm', '.xls')):
                self.log("Geçersiz dosya formatı! (.xlsx, .xlsm veya .xls olmalı)", "error")
                APP_LOGGER.error(f"Geçersiz format: {path}")
                return False
            
            # Dosya boyutu kontrolü (max 100MB)
            file_size = os.path.getsize(path)
            max_size = 100 * 1024 * 1024  # 100 MB
            if file_size > max_size:
                size_mb = file_size / (1024 * 1024)
                self.log(f"Dosya çok büyük! ({size_mb:.1f}MB - Max: 100MB)", "error")
                APP_LOGGER.error(f"Dosya çok büyük: {size_mb:.1f}MB")
                return False
            
            # Dosya okuma izni kontrolü
            if not os.access(path, os.R_OK):
                self.log("Dosya okunamıyor! İzin hatası.", "error")
                APP_LOGGER.error(f"Okuma izni yok: {path}")
                return False
            
            APP_LOGGER.info(f"Dosya validasyonu başarılı: {path} ({file_size / 1024:.1f}KB)")
            return True
            
        except Exception as e:
            self.log(f"Dosya validasyon hatası: {e}", "error")
            APP_LOGGER.error(f"Validasyon hatası: {e}\n{traceback.format_exc()}")
            return False

    def update_ui_with_excel_data(self, sheets, data):
        # Sayfa listesini güncelle
        self.combo_sheet.configure(values=sheets)
        if sheets:
            self.combo_sheet.set(sheets[0])
            self.on_sheet_change(self.combo_sheet.get())

        if self.empty_state_frame is not None:
            self.empty_state_frame.destroy()
            self.empty_state_frame = None
        if self.preview_table is not None:
            self.preview_table.destroy()
            self.preview_table = None
            
        if not data:
            self.preview_table = ctk.CTkFrame(self.preview_box, fg_color="transparent")
            self.preview_table.pack(fill="both", expand=True, padx=15, pady=15)
            ctk.CTkLabel(self.preview_table, text="Veri okunamadı.", text_color="red").pack()
            return
        
        self.preview_table = ctk.CTkFrame(self.preview_box, fg_color="transparent")
        self.preview_table.pack(fill="both", expand=True, padx=15, pady=15)
        num_cols = len(data[0]) if data else 0
        for idx in range(num_cols + 1):
            self.preview_table.grid_columnconfigure(idx, weight=1)

        # Satır numarası başlığı
        row_header = ctk.CTkLabel(self.preview_table, text="#", width=35,
                                 fg_color=self.colors["grid_header"],
                                 font=("Roboto", 10, "bold"), height=30, corner_radius=4)
        row_header.grid(row=0, column=0, padx=1, pady=(0,2), sticky="ew")
        
        for c_idx in range(num_cols):
            col_letter = num2col(c_idx + 1)
            width = 75
            header = ctk.CTkLabel(self.preview_table, text=col_letter, 
                                  fg_color=self.colors["grid_header"], 
                                  font=("Roboto", 11, "bold"), height=30, 
                                  corner_radius=4, width=width)
            header.grid(row=0, column=c_idx+1, padx=1, pady=(0,2), sticky="ew")

        for r_idx, row_data in enumerate(data):
            # Satır numarası
            row_num = ctk.CTkLabel(self.preview_table, text=str(r_idx+1), width=35, 
                                   text_color="#aaa", font=("Arial", 9),
                                   fg_color=self.colors["grid_row_even"] if r_idx % 2 == 0 else self.colors["grid_row_odd"])
            row_num.grid(row=r_idx+1, column=0, padx=1, pady=1, sticky="ew")
            
            bg_color = self.colors["grid_row_even"] if r_idx % 2 == 0 else self.colors["grid_row_odd"]
            
            for c_idx, cell_val in enumerate(row_data):
                val_str = str(cell_val) if cell_val is not None else ""
                if len(val_str) > 10: 
                    val_str = val_str[:8] + "..."
                
                width = 75
                cell = ctk.CTkLabel(self.preview_table, text=val_str, height=26,
                                    fg_color=bg_color, font=("Consolas", 10), 
                                    width=width, anchor="w", padx=5)
                cell.grid(row=r_idx+1, column=c_idx+1, padx=1, pady=1, sticky="ew")
        self.log("Önizleme yüklendi.", "success")

    def start_process(self):
        """İşlemi başlat - Gelişmiş validasyon ile"""
        try:
            # Parametreleri topla
            dynamic_params = []
            for row in self.param_rows:
                suffix = row["name"].get().strip()
                col = row["col"].get().strip()
                if col:
                    # Sütun validasyonu
                    if not col.isalpha():
                        self.log(f"Geçersiz sütun: '{col}' - Sadece harf kullanın (A-Z)", "error")
                        return
                    dynamic_params.append((suffix, col))
            
            # Temel validasyonlar
            if not self.selected_file:
                self.log("Önce bir Excel dosyası seçmelisiniz.", "error")
                messagebox.showwarning("Uyarı", "Lütfen önce bir Excel dosyası seçin!")
                return
            
            if not dynamic_params:
                self.log("En az bir parametre eşleştirmesi yapmalısınız!", "error")
                messagebox.showwarning("Uyarı", "Ayarlar sekmesinden parametre eşleştirmesi yapın!")
                return
            
            # Dosya tekrar validate et
            if not self.validate_excel_file(self.selected_file):
                self.log("Dosya validasyonu başarısız!", "error")
                return
            
            self.config["sheet_name"] = self.combo_sheet.get()
            
            if not self.config["sheet_name"]:
                self.log("Sayfa seçilmedi!", "error")
                messagebox.showwarning("Uyarı", "Lütfen bir Excel sayfası seçin!")
                return
            
            # Kullanıcıdan onay al (opsiyonel)
            confirm_msg = f"İşlem başlatılacak:\n\n"
            confirm_msg += f"Dosya: {os.path.basename(self.selected_file)}\n"
            confirm_msg += f"Sayfa: {self.config['sheet_name']}\n"
            confirm_msg += f"Parametre Sayısı: {len(dynamic_params)}\n\n"
            confirm_msg += "Devam etmek istiyor musunuz?"
            
            if not messagebox.askyesno("İşlemi Başlat", confirm_msg):
                self.log("İşlem kullanıcı tarafından iptal edildi.", "info")
                return
            
            # UI'yi hazırla
            self.tab_view.set("  Monitör  ")
            self.set_controls_state(False)
            self.set_running_state(True)
            self.btn_run.configure(state="disabled", text="ÇALIŞIYOR...", fg_color="#555555")
            self.log_entries.clear()
            self.render_log()
            self.current_run_has_error = False
            self.btn_stop.configure(state="normal")
            
            # İşlemi başlat
            self.start_time = time.time()
            APP_LOGGER.info(f"İşlem başlatılıyor - Dosya: {self.selected_file}, Sayfa: {self.config['sheet_name']}")
            APP_LOGGER.info(f"Parametreler: {dynamic_params}")
            
            self.worker = WorkerThread(self, self.selected_file, self.config, dynamic_params)
            self.worker.start()
            
        except Exception as e:
            APP_LOGGER.error(f"İşlem başlatma hatası: {e}\n{traceback.format_exc()}")
            self.log(f"İşlem başlatılamadı: {e}", "error")
            messagebox.showerror("Hata", f"İşlem başlatılamadı:\n{str(e)}")

    # (Diğer yardımcı fonksiyonlar: log, update_stats vs. aynı kalır)
    def log(self, msg, type="info"):
        icons = {"info": "ℹ", "update": "⚡", "error": "✖", "success": "✔"}
        ts = datetime.datetime.now().strftime("%H:%M:%S")
        self.log_entries.append({"ts": ts, "type": type, "icon": icons.get(type, ""), "msg": msg})
        
        # Log sayısını sınırla (bellek optimizasyonu)
        max_logs = 5000
        if len(self.log_entries) > max_logs:
            self.log_entries = self.log_entries[-max_logs:]
        
        if type == "error":
            self.current_run_has_error = True
            if self.btn_run.cget("state") == "normal":
                self.btn_run.configure(fg_color=self.colors["danger"])
        elif type == "success" and self.btn_run.cget("state") == "normal":
            self.btn_run.configure(fg_color="#2ECC71")
        
        # Throttle log rendering (her 10 log'da bir render et, hata/başarı için hemen)
        if not hasattr(self, '_log_render_count'):
            self._log_render_count = 0
        self._log_render_count += 1
        if self._log_render_count % 10 == 0 or type in ["error", "success"]:
            self.render_log()

    def render_log(self):
        """Log render'ını optimize et - sadece son 1000 satırı göster"""
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        
        # Filtreleme
        filtered = self.log_entries
        if self.show_errors_only:
            filtered = [e for e in self.log_entries if e["type"] == "error"]
        
        # Performans: Çok fazla log varsa sadece son 1000'ini göster
        max_display = 1000
        if len(filtered) > max_display:
            filtered = filtered[-max_display:]
            self.log_box.insert("end", f"... ({len(self.log_entries) - max_display} eski log gizlendi) ...\n")
        
        # Batch insert (daha hızlı)
        log_text = "\n".join([f"[{e['ts']}] {e['icon']} {e['msg']}" for e in filtered])
        self.log_box.insert("end", log_text + "\n")
        self.log_box.see("end")
        self.log_box.configure(state="disabled")
        self.update_error_button_text()

    def toggle_error_filter(self):
        self.show_errors_only = not self.show_errors_only
        self.render_log()

    def save_log(self):
        if not self.log_entries:
            self.log("Kaydedilecek log bulunamadı.", "error")
            return
        if not os.path.exists("Logs"):
            os.makedirs("Logs")
        filename = os.path.join("Logs", f"log_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.txt")
        with open(filename, "w", encoding="utf-8") as f:
            for entry in self.log_entries:
                f.write(f"[{entry['ts']}] {entry['icon']} {entry['msg']}\n")
        self.log(f"Log kaydedildi: {filename}", "success")
        
    def update_stats(self, done, updates, errors):
        ratio = done / self.total_work if self.total_work > 0 else 0
        self.progress.set(ratio)
        self.lbl_upd.configure(text=str(updates))
        self.lbl_err.configure(text=str(errors))
        elapsed = time.time() - self.start_time
        if elapsed > 0:
            speed = done/elapsed if elapsed > 0 else 0.001
            rem = (self.total_work - done) / speed
            self.lbl_time.configure(text=time.strftime('%M:%S', time.gmtime(elapsed)))
            self.lbl_eta.configure(text=time.strftime('%M:%S', time.gmtime(rem)))
            
            # Gelişmiş istatistikler (başarı oranı)
            if done > 0:
                success_rate = (updates / done) * 100
                if hasattr(self, 'lbl_success_rate'):
                    self.lbl_success_rate.configure(text=f"{success_rate:.1f}%")
                if hasattr(self, 'lbl_avg_time'):
                    avg_time = elapsed / done if done > 0 else 0
                    self.lbl_avg_time.configure(text=f"{avg_time:.2f}s")

    def update_max_progress(self, val): 
        self.total_work = val
        self.start_time = time.time()
        self.progress.set(0)
        self.start_hamster_animation()
        
    def initial_draw_hamster(self):
        """İlk hamster çizimi"""
        if hasattr(self, 'hamster_canvas'):
            self.hamster_canvas.update_idletasks()
            self.draw_hamster_wheel()
        
    def draw_hamster_wheel(self):
        """Hamster çarkı ve hamster'ı çizer"""
        if not hasattr(self, 'hamster_canvas'):
            return
        canvas = self.hamster_canvas
        canvas.delete("all")
        
        # Canvas boyutları
        width = canvas.winfo_width()
        height = canvas.winfo_height()
        if width < 10 or height < 10:
            # Minimum boyut ayarla
            width = max(width, 200)
            height = max(height, 80)
            canvas.configure(width=width, height=height)
        
        center_x = width // 2
        center_y = height // 2
        wheel_radius = min(width, height) // 3
        
        # Çark (dış halka)
        canvas.create_oval(
            center_x - wheel_radius, center_y - wheel_radius,
            center_x + wheel_radius, center_y + wheel_radius,
            outline="#666", width=3, fill="#2b2b2b"
        )
        
        # İç halka
        inner_radius = wheel_radius * 0.7
        canvas.create_oval(
            center_x - inner_radius, center_y - inner_radius,
            center_x + inner_radius, center_y + inner_radius,
            outline="#444", width=2, fill="#1a1a1a"
        )
        
        # Spokes (çubuklar)
        import math
        num_spokes = 8
        for i in range(num_spokes):
            angle = (self.hamster_angle + i * 360 / num_spokes) * math.pi / 180
            x1 = center_x + inner_radius * math.cos(angle)
            y1 = center_y + inner_radius * math.sin(angle)
            x2 = center_x + wheel_radius * math.cos(angle)
            y2 = center_y + wheel_radius * math.sin(angle)
            canvas.create_line(x1, y1, x2, y2, fill="#555", width=2)
        
        # Hamster (basitleştirilmiş - daire)
        hamster_angle_rad = (self.hamster_angle + 90) * math.pi / 180
        hamster_x = center_x + (inner_radius + wheel_radius) / 2 * math.cos(hamster_angle_rad)
        hamster_y = center_y + (inner_radius + wheel_radius) / 2 * math.sin(hamster_angle_rad)
        hamster_size = wheel_radius * 0.15
        
        # Hamster gövdesi
        canvas.create_oval(
            hamster_x - hamster_size, hamster_y - hamster_size,
            hamster_x + hamster_size, hamster_y + hamster_size,
            fill="#D2691E", outline="#8B4513", width=2
        )
        
        # Hamster gözü
        eye_offset = hamster_size * 0.3
        eye_size = hamster_size * 0.2
        eye_x = hamster_x + eye_offset * math.cos(hamster_angle_rad)
        eye_y = hamster_y + eye_offset * math.sin(hamster_angle_rad)
        canvas.create_oval(
            eye_x - eye_size, eye_y - eye_size,
            eye_x + eye_size, eye_y + eye_size,
            fill="white", outline="black"
        )
        canvas.create_oval(
            eye_x - eye_size*0.5, eye_y - eye_size*0.5,
            eye_x + eye_size*0.5, eye_y + eye_size*0.5,
            fill="black"
        )
        
    def animate_hamster(self):
        """Hamster animasyonunu günceller"""
        if not hasattr(self, 'hamster_animation_id') or self.hamster_animation_id is None:
            return
        
        self.hamster_angle = (self.hamster_angle + 8) % 360
        self.draw_hamster_wheel()
        
        # Animasyonu devam ettir (sadece çalışıyorsa)
        if self.hamster_animation_id is not None:
            self.hamster_animation_id = self.after(50, self.animate_hamster)
        
    def start_hamster_animation(self):
        """Hamster animasyonunu başlatır"""
        # Eğer zaten çalışıyorsa başlatma
        if hasattr(self, 'hamster_animation_id') and self.hamster_animation_id is not None:
            return
        
        # Canvas'ın hazır olduğundan emin ol
        if hasattr(self, 'hamster_canvas'):
            self.hamster_canvas.update_idletasks()
            
        self.hamster_angle = 0
        self.draw_hamster_wheel()
        # İlk animasyon çağrısını başlat
        self.hamster_animation_id = self.after(50, self.animate_hamster)
            
    def stop_hamster_animation(self):
        """Hamster animasyonunu durdurur"""
        if self.hamster_animation_id is not None:
            self.after_cancel(self.hamster_animation_id)
            self.hamster_animation_id = None
        
    def finish_process(self):
        self.set_controls_state(True)
        self.btn_stop.configure(state="disabled")
        self.set_running_state(False)
        self.stop_hamster_animation()
        fg = self.colors["danger"] if self.current_run_has_error else "#2ECC71"
        self.btn_run.configure(state="normal", text="İŞLEMİ BAŞLAT ▶", fg_color=fg)
        
        # İşlem özeti
        elapsed = time.time() - self.start_time
        total_updates = int(self.lbl_upd.cget("text") or 0)
        total_errors = int(self.lbl_err.cget("text") or 0)
        success_rate = (total_updates / self.total_work * 100) if self.total_work > 0 else 0
        
        APP_LOGGER.info(f"İşlem tamamlandı - Başarılı: {total_updates}, Hata: {total_errors}, Oran: {success_rate:.1f}%")
        
        self.log("İşlem tamamlandı.", "success")
        self.log(f"Özet: {total_updates} başarılı, {total_errors} hata, {success_rate:.1f}% başarı oranı", "info")
        
        # Otomatik kayıt (her zaman kaydet)
        try:
            self.auto_save_results(total_updates, total_errors, elapsed)
        except Exception as e:
            APP_LOGGER.error(f"Otomatik kayıt hatası: {e}")
        
        # Bildirim göster (eğer hatalar varsa)
        if total_errors > 0:
            self.log(f"⚠️ {total_errors} hata oluştu! Detaylar için logu kontrol edin.", "error")
    
    def auto_save_results(self, updates, errors, elapsed_time):
        """İşlem sonuçlarını otomatik kaydeder"""
        if not os.path.exists("Results"):
            os.makedirs("Results")
        
        filename = os.path.join("Results", f"result_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.txt")
        try:
            with open(filename, "w", encoding="utf-8") as f:
                f.write(f"CATIA Automation Suite - İşlem Sonuçları\n")
                f.write(f"{'='*50}\n\n")
                f.write(f"Tarih: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Dosya: {os.path.basename(self.selected_file) if self.selected_file else 'N/A'}\n")
                f.write(f"Sayfa: {self.config.get('sheet_name', 'N/A')}\n\n")
                f.write(f"Toplam İşlenen: {self.total_work}\n")
                f.write(f"Başarılı: {updates}\n")
                f.write(f"Hatalar: {errors}\n")
                f.write(f"Başarı Oranı: {(updates/self.total_work*100) if self.total_work > 0 else 0:.1f}%\n")
                f.write(f"Geçen Süre: {time.strftime('%M:%S', time.gmtime(elapsed_time))}\n")
            self.log(f"Sonuçlar otomatik kaydedildi: {os.path.basename(filename)}", "info")
        except Exception as e:
            self.log(f"Sonuç kaydedilemedi: {e}", "error")
    
    # ------------------------------
    # PROFİL YÖNETİMİ (Kaydet/Yükle)
    # ------------------------------
    def save_profile(self):
        """Mevcut ayarları JSON dosyasına kaydeder"""
        profile = {
            "params": [],
            "sheet_name": self.config.get("sheet_name", ""),
            "excel_file": self.selected_file if self.selected_file else ""
        }
        for row in self.param_rows:
            suffix = row["name"].get().strip()
            col = row["col"].get().strip()
            if suffix or col:
                profile["params"].append({"suffix": suffix, "col": col})
        
        filename = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON dosyaları", "*.json"), ("Tüm dosyalar", "*.*")],
            title="Profil Kaydet"
        )
        if filename:
            try:
                with open(filename, "w", encoding="utf-8") as f:
                    json.dump(profile, f, indent=2, ensure_ascii=False)
                self.log(f"Profil kaydedildi: {os.path.basename(filename)}", "success")
            except Exception as e:
                self.log(f"Profil kaydedilemedi: {e}", "error")
    
    def load_profile(self):
        """JSON dosyasından ayarları yükler"""
        filename = filedialog.askopenfilename(
            filetypes=[("JSON dosyaları", "*.json"), ("Tüm dosyalar", "*.*")],
            title="Profil Yükle"
        )
        if filename:
            try:
                with open(filename, "r", encoding="utf-8") as f:
                    profile = json.load(f)
                
                # Mevcut parametreleri temizle
                for row in self.param_rows:
                    row["frame"].destroy()
                self.param_rows = []
                
                # Yeni parametreleri ekle
                if "params" in profile:
                    for p in profile["params"]:
                        self.add_param_row(p.get("suffix", ""), p.get("col", ""))
                
                # Sayfa adını güncelle
                if "sheet_name" in profile and profile["sheet_name"]:
                    self.config["sheet_name"] = profile["sheet_name"]
                    if hasattr(self, "combo_sheet"):
                        try:
                            self.combo_sheet.set(profile["sheet_name"])
                        except:
                            pass
                
                # Excel dosyasını yükle (eğer varsa)
                if "excel_file" in profile and profile["excel_file"] and os.path.exists(profile["excel_file"]):
                    self.selected_file = profile["excel_file"]
                    self.update_summary_label()
                    loader = ExcelPreviewLoader(self, profile["excel_file"])
                    loader.start()
                
                self.log(f"Profil yüklendi: {os.path.basename(filename)}", "success")
            except Exception as e:
                self.log(f"Profil yüklenemedi: {e}", "error")
    
    def load_default_profile(self):
        """Varsayılan profili yüklemeyi dene"""
        default_path = "default_profile.json"
        if os.path.exists(default_path):
            try:
                with open(default_path, "r", encoding="utf-8") as f:
                    profile = json.load(f)
                if "params" in profile:
                    for row in self.param_rows:
                        row["frame"].destroy()
                    self.param_rows = []
                    for p in profile["params"]:
                        self.add_param_row(p.get("suffix", ""), p.get("col", ""))
            except:
                pass
    
    # ------------------------------
    # PARAMETRE ŞABLONLARI
    # ------------------------------
    def show_template_menu(self):
        """Parametre şablonları menüsünü gösterir"""
        templates = {
            "Standart": [("Thickness", "B"), ("H", "C"), ("P1", "K"), ("D1", "L"), ("P2", "M"), ("D2", "N")],
            "Basit": [("Thickness", "B"), ("H", "C")],
            "Detaylı": [("Thickness", "B"), ("H", "C"), ("P1", "K"), ("D1", "L"), ("P2", "M"), ("D2", "N"), ("P3", "O"), ("D3", "P")],
            "Özel 1": [("T", "B"), ("W", "C"), ("L", "D")],
            "Özel 2": [("Height", "B"), ("Width", "C"), ("Depth", "D"), ("Angle", "E")]
        }
        
        # Popup pencere
        popup = ctk.CTkToplevel(self)
        popup.title("Parametre Şablonları")
        popup.geometry("400x350")
        popup.transient(self)
        popup.grab_set()
        
        ctk.CTkLabel(popup, text="Bir şablon seçin:", font=("Roboto", 14, "bold")).pack(pady=15)
        
        scroll = ctk.CTkScrollableFrame(popup)
        scroll.pack(fill="both", expand=True, padx=20, pady=10)
        
        for name, params in templates.items():
            btn = ctk.CTkButton(
                scroll,
                text=f"📋 {name} ({len(params)} parametre)",
                command=lambda p=params: self.apply_template(p, popup),
                fg_color=self.colors["panel"],
                hover_color=self.colors["card"],
                height=40,
                anchor="w"
            )
            btn.pack(fill="x", pady=5)
        
        ctk.CTkButton(
            popup,
            text="İptal",
            command=popup.destroy,
            fg_color=self.colors["danger"],
            height=35
        ).pack(pady=10)
    
    def apply_template(self, params, popup):
        """Şablonu uygular"""
        # Mevcut parametreleri temizle
        for row in self.param_rows:
            row["frame"].destroy()
        self.param_rows = []
        
        # Yeni parametreleri ekle
        for suffix, col in params:
            self.add_param_row(suffix, col)
        
        popup.destroy()
        self.log(f"Şablon uygulandı ({len(params)} parametre)", "success")
        self.btn_stop.configure(state="disabled")
        self.set_running_state(False)
        self.stop_hamster_animation()
        fg = self.colors["danger"] if self.current_run_has_error else "#2ECC71"
        self.btn_run.configure(state="normal", text="İŞLEMİ BAŞLAT ▶", fg_color=fg)
        self.log("İşlem tamamlandı.", "success")
        
    def stop_process(self):
        if self.worker:
            self.worker.stop()
            self.btn_stop.configure(state="disabled")
            self.status_badge.configure(text="Durum: Durduruluyor", text_color="#F39C12")
            self.log("Durdurma talebi gönderildi.", "info")
            self.stop_hamster_animation()
    
    # ------------------------------
    # YENİ ÖZELLİKLER
    # ------------------------------
    def create_excel_template(self):
        """Excel şablonu oluştur"""
        try:
            if not OPENPYXL_AVAILABLE:
                messagebox.showerror("Hata", "openpyxl yüklü değil!\n'pip install openpyxl' ile yükleyin.")
                return
            
            # Parametreleri topla
            params = []
            for row in self.param_rows:
                suffix = row["name"].get().strip()
                col = row["col"].get().strip()
                if suffix:
                    params.append((suffix, col))
            
            if not params:
                messagebox.showwarning("Uyarı", "Önce parametre ekleyin!")
                return
            
            # Dosya adı sor
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel Dosyaları", "*.xlsx"), ("Tüm Dosyalar", "*.*")],
                title="Excel Şablonu Kaydet",
                initialfile=f"CATIA_Template_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx"
            )
            
            if filename:
                generate_excel_template(params, filename)
                self.log(f"Şablon oluşturuldu: {os.path.basename(filename)}", "success")
                messagebox.showinfo("Başarılı", f"Excel şablonu oluşturuldu:\n{os.path.basename(filename)}")
                
        except Exception as e:
            APP_LOGGER.error(f"Şablon oluşturma hatası: {e}\n{traceback.format_exc()}")
            self.log(f"Şablon oluşturulamadı: {e}", "error")
            messagebox.showerror("Hata", f"Şablon oluşturulamadı:\n{str(e)}")
    
    def export_html_report(self):
        """HTML raporu export et"""
        try:
            if not self.log_entries:
                messagebox.showwarning("Uyarı", "Henüz log verisi yok!")
                return
            
            # İstatistikleri topla
            total = int(self.lbl_upd.cget("text") or 0) + int(self.lbl_err.cget("text") or 0)
            updates = int(self.lbl_upd.cget("text") or 0)
            errors = int(self.lbl_err.cget("text") or 0)
            success_rate = (updates / total * 100) if total > 0 else 0
            
            elapsed = time.time() - self.start_time if hasattr(self, 'start_time') else 0
            elapsed_str = time.strftime('%M:%S', time.gmtime(elapsed))
            
            stats = {
                'total': total,
                'updates': updates,
                'errors': errors,
                'success_rate': success_rate,
                'elapsed_time': elapsed_str,
                'filename': os.path.basename(self.selected_file) if self.selected_file else "N/A",
                'sheet': self.config.get('sheet_name', 'N/A')
            }
            
            # Dosya adı sor
            if not os.path.exists("Reports"):
                os.makedirs("Reports")
            
            filename = filedialog.asksaveasfilename(
                defaultextension=".html",
                filetypes=[("HTML Dosyaları", "*.html"), ("Tüm Dosyalar", "*.*")],
                title="Raporu Kaydet",
                initialdir="Reports",
                initialfile=f"Report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
            )
            
            if filename:
                generate_html_report(self.log_entries, stats, filename)
                self.log(f"HTML raporu oluşturuldu: {os.path.basename(filename)}", "success")
                
                # Raporu aç?
                if messagebox.askyesno("Başarılı", "Rapor oluşturuldu!\n\nBrowser'da açmak ister misiniz?"):
                    import webbrowser
                    webbrowser.open(f"file://{os.path.abspath(filename)}")
                
        except Exception as e:
            APP_LOGGER.error(f"Rapor oluşturma hatası: {e}\n{traceback.format_exc()}")
            self.log(f"Rapor oluşturulamadı: {e}", "error")
            messagebox.showerror("Hata", f"Rapor oluşturulamadı:\n{str(e)}")
    
    def on_closing(self):
        """Uygulama kapatılırken"""
        APP_LOGGER.info("Uygulama kapatılıyor...")
        
        # Çalışan işlem varsa sor
        if self.worker and self.worker.is_alive():
            if messagebox.askyesnocancel("Çıkış", "İşlem devam ediyor!\n\nYine de çıkmak istiyor musunuz?"):
                self.worker.stop()
                self.worker.join(timeout=2)
                self.destroy()
        else:
            self.destroy()

if __name__ == "__main__":
    APP_LOGGER.info("=" * 50)
    APP_LOGGER.info("CATIA Automation Suite v4.5 Pro - Başlatılıyor")
    APP_LOGGER.info(f"Test Modu: {TEST_MODE}")
    APP_LOGGER.info(f"openpyxl: {'Yüklü' if OPENPYXL_AVAILABLE else 'Yüklü Değil'}")
    APP_LOGGER.info(f"win32com: {'Yüklü' if WIN32COM_AVAILABLE else 'Yüklü Değil'}")
    APP_LOGGER.info("=" * 50)
    
    try:
        app = AutomationSuite()
        app.mainloop()
    except Exception as e:
        APP_LOGGER.critical(f"Kritik uygulama hatası: {e}\n{traceback.format_exc()}")
        messagebox.showerror("Kritik Hata", f"Uygulama başlatılamadı:\n{str(e)}")
    finally:
        APP_LOGGER.info("Uygulama kapatıldı")